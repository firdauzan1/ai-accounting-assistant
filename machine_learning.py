# In[1]:

# Cell 1 - Basic Imports
import os
import json
import logging
import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional
from datetime import datetime
import difflib

# Konfigurasi Logging (jika belum ada)
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Import modul untuk ekspor dan impor Excel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# In[2]:

# Cell 1.2 - ML & NLP Setup
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.neighbors import NearestNeighbors
from transformers import pipeline
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier 
from sklearn.naive_bayes import MultinomialNB 


# NLP Setup: Load model spaCy
import spacy
try:
    nlp_spacy = spacy.load("en_core_web_sm")
except OSError:
    from spacy.cli import download
    download("en_core_web_sm")
    nlp_spacy = spacy.load("en_core_web_sm")

# Handle PyTorch compatibility jika diperlukan
import torch
if hasattr(torch.utils._pytree, 'register_pytree_node'):
    from transformers.utils import generic
    generic._torch_pytree._register_pytree_node = torch.utils._pytree.register_pytree_node

# In[3]:

# Cell 1.3 - Translator & Logging Setup
import asyncio
from googletrans import Translator

# Initialize the translator
translator = Translator()

def translate_input(text: str) -> str:
    """
    Menerjemahkan teks dari Bahasa Indonesia ke Bahasa Inggris untuk proses model.
    Parameters:
        text (str): Teks dalam Bahasa Indonesia.
    Returns:
        str: Teks yang telah diterjemahkan ke Bahasa Inggris.
    """
    try:
        if text and any(char.isalpha() for char in text):
            # Gunakan metode sinkron dengan asyncio
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            translation = translator.translate(text, src='id', dest='en')
            loop.close()
            logger.info(f"Terjemahan: {text} -> {translation.text}")
            return translation.text
        return text
    except Exception as e:
        logger.error(f"Gagal menerjemahkan: {str(e)}")
        return text

# In[4]:

# Cell 1.4 - File Configuration & Constants
# File Configuration
FILES = {
    'chart': 'data/chart_of_account.xlsx',
    'positions': 'data/normal_account_position.xlsx',
    'transactions': 'data/transactions.json',
    'feedback': 'data/transactions_feedback.json'
}

FEEDBACK_MULTIPLIER = 3
ML_CONFIDENCE_THRESHOLD = 0.4

# Definisi warna untuk masing-masing tipe akun (digunakan pada buku besar)
ACCOUNT_COLORS = {
    'asset': '#E3F2FD',      # Biru muda - tetap
    'liability': '#FFF3E0',  # Oranye muda - tetap
    'equity': '#E8F5E9',     # Hijau muda - tetap
    'revenue': '#E0E0E0',    # Ganti ungu muda (#F3E5F5) dengan abu-abu muda
    'expense': '#F5F5F5'     # Ganti merah muda (#FFEBEE) dengan abu-abu sangat muda
}


# In[5]:

# Cell 2 - LoginWindow Base Class
class LoginWindow:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("AI Accounting Assistant") 
        self.window.geometry("640x480")  # Perbesar ukuran window untuk tombol nama perusahaan
        self.window.configure(bg='#283e4a')
        
        # Font definitions
        self.title_font = ('Helvetica', 16, 'bold')
        self.label_font = ('Helvetica', 12)
        self.entry_font = ('Helvetica', 12)
        self.btn_font = ('Helvetica', 12, 'bold')
        
        # --- TAMBAHKAN KODE INI UNTUK FAVICON ---
        try:
            # Pastikan logo.png ada di direktori yang sama dengan script Python Anda
            icon = tk.PhotoImage(file='assets/logo.png')
            # Argumen False berarti ikon ini berlaku untuk jendela ini dan turunannya
            self.window.iconphoto(False, icon) 
            logger.info("Favicon logo.png berhasil dimuat untuk LoginWindow.")
        except tk.TclError as e:
            # Error ini sering terjadi jika file tidak ditemukan atau format tidak didukung
            logger.warning(f"Gagal memuat favicon logo.png untuk LoginWindow: {e}. Pastikan file 'logo.png' ada di direktori yang sama dan formatnya didukung (GIF/PGM/PPM atau PNG jika versi Tkinter Anda mendukungnya).")
        except Exception as e:
             # Menangkap error tak terduga lainnya
             logger.warning(f"Terjadi kesalahan tak terduga saat memuat favicon logo.png untuk LoginWindow: {e}")
        # --- AKHIR KODE FAVICON ---
        
        # Initialize tracking variables
        self.success = False
        self.company_name = "PT ABCD"  # Default company name


# In[6]:


# Cell 2.1 - LoginWindow UI Setup
def setup_ui(self):
    # Main container frame
    self.main_frame = tk.Frame(self.window, bg='#ffffff', bd=2, relief='raised')
    self.main_frame.place(relx=0.5, rely=0.5, anchor='center', width=350, height=250)
    
    # Title label
    self.title_label = tk.Label(self.main_frame, text="Selamat Datang", font=self.title_font, bg='#ffffff', fg='#283e4a')
    self.title_label.pack(pady=(10,10))
    
    # Company Name input
    self.company_frame = tk.Frame(self.main_frame, bg='#ffffff')
    self.company_frame.pack(fill='x', pady=5, padx=10)
    self.company_label = tk.Label(self.company_frame, text="Nama Perusahaan:", font=self.label_font, bg='#ffffff', fg='#283e4a')
    self.company_label.pack(side='left')
    self.company_var = tk.StringVar(value="PT ABCD")
    self.company_entry = tk.Entry(self.company_frame, textvariable=self.company_var, font=self.entry_font, relief='solid', bd=1)
    self.company_entry.pack(side='right', expand=True, fill='x')
    
    # Username input
    self.username_frame = tk.Frame(self.main_frame, bg='#ffffff')
    self.username_frame.pack(fill='x', pady=5, padx=10)
    self.username_label = tk.Label(self.username_frame, text="Nama Pengguna:", font=self.label_font, bg='#ffffff', fg='#283e4a')
    self.username_label.pack(side='left')
    self.username_var = tk.StringVar()
    self.username_entry = tk.Entry(self.username_frame, textvariable=self.username_var, font=self.entry_font, relief='solid', bd=1)
    self.username_entry.pack(side='right', expand=True, fill='x')
    
    # Password input
    self.password_frame = tk.Frame(self.main_frame, bg='#ffffff')
    self.password_frame.pack(fill='x', pady=5, padx=10)
    self.password_label = tk.Label(self.password_frame, text="Kata Sandi:", font=self.label_font, bg='#ffffff', fg='#283e4a')
    self.password_label.pack(side='left')
    self.password_var = tk.StringVar()
    self.password_entry = tk.Entry(self.password_frame, textvariable=self.password_var, show="•", font=self.entry_font, relief='solid', bd=1)
    self.password_entry.pack(side='right', expand=True, fill='x')
    
    # Error message
    self.error_var = tk.StringVar()
    self.error_label = tk.Label(self.main_frame, textvariable=self.error_var, font=('Helvetica', 10), fg='red', bg='#ffffff')
    self.error_label.pack(pady=(5,0))
    
    # Login button
    self.login_button = tk.Button(self.main_frame, text="Masuk", command=self.login,
                               font=self.btn_font, bg='#4CAF50', fg='white',
                               bd=0, activebackground='#45a049', cursor='hand2')
    self.login_button.pack(pady=(15,10))
    
    # Event bindings
    self.window.bind('<Return>', lambda e: self.login())
    self.company_entry.focus()

# Add the method to LoginWindow
LoginWindow.setup_ui = setup_ui


# In[7]:


# Cell 2.2 - LoginWindow Authentication
def login(self):
    username = self.username_var.get().strip()
    password = self.password_var.get().strip()
    
    if username == "admin" and password == "admin":
        self.success = True
        self.company_name = self.company_var.get().strip()
        if not self.company_name:
            self.company_name = "PT ABCD"
        self.window.destroy()
    else:
        self.error_var.set("Nama pengguna atau kata sandi salah")
        self.password_var.set("")
        self.password_entry.focus()

def run(self):
    # Setup the UI elements before running
    self.setup_ui()
    self.window.mainloop()
    return self.success, self.company_name  # Mengembalikan tuple (success, company_name)

# Add the methods to LoginWindow
LoginWindow.login = login
LoginWindow.run = run


# In[8]:


# Cell 3 - AccountingAssistant Kelas Dasar
@dataclass
class AccountRule:
    increase: str
    decrease: str
    description: str
    debit_rule: str
    credit_rule: str

class AccountingAssistant:
    def __init__(self):
        # Inisialisasi semua atribut terkait model ke None/False
        self.vectorizer = None
        self.nn = None
        self.classifier_available = False
        self.debit_rf_classifier = None
        self.credit_rf_classifier = None
        self.debit_lr_classifier = None
        self.credit_lr_classifier = None
        self.debit_nb_classifier = None
        self.credit_nb_classifier = None

        # Muat data
        self.account_data = self._load_account_data()
        self.transactions = self._load_transactions()
        self.feedback_data = self._load_feedback_data()

        boosted_feedback = self.feedback_data * FEEDBACK_MULTIPLIER
        self.training_data = self.transactions + boosted_feedback

        if self.training_data:
            # Coba inisialisasi model
            try:
                self._init_models()  # _init_models akan mencoba set classifier_available=True
            except Exception as e:
                logger.error(f"Gagal menginisialisasi model di __init__: {e}")
                # Pastikan classifier_available tetap False jika _init_models gagal
                self.classifier_available = False
        else:
            logger.error("Tidak ada data training yang berhasil dimuat. Model tidak dapat diinisialisasi.")
            # Atribut sudah diinisialisasi ke None/False di atas

    def _load_account_data(self) -> Dict[str, dict]:
        logger.info("Attempting to load account data from Excel files...")
        try:
            chart_path = FILES.get('chart', 'chart_of_account.xlsx')
            positions_path = FILES.get('positions', 'normal_account_position.xlsx')
            logger.info(f"Chart path: {chart_path}")
            logger.info(f"Positions path: {positions_path}")

            if not os.path.exists(chart_path):
                logger.error(f"CRITICAL ERROR: Chart of Account file NOT FOUND at '{chart_path}'.")
                messagebox.showerror("File Error", f"File Chart of Account tidak ditemukan:\n{chart_path}")
                return {}
            if not os.path.exists(positions_path):
                logger.error(f"CRITICAL ERROR: Normal Account Position file NOT FOUND at '{positions_path}'.")
                messagebox.showerror("File Error", f"File Normal Account Position tidak ditemukan:\n{positions_path}")
                return {}

            logger.info("Reading Excel files...")
            # Explicitly set dtype to string for potentially problematic columns to avoid inference issues
            try:
                chart_df = pd.read_excel(chart_path, sheet_name=0, dtype={'account_name': str, 'account_chart': str})
                pos_df = pd.read_excel(positions_path, sheet_name=0, dtype={'account_name': str, 'normal_account_position': str, 'description': str})
            except Exception as e:
                 logger.error(f"Error reading Excel files with specified dtypes: {e}", exc_info=True)
                 # Fallback to default reading if dtype specification fails
                 logger.warning("Falling back to reading Excel without explicit dtypes.")
                 chart_df = pd.read_excel(chart_path, sheet_name=0)
                 pos_df = pd.read_excel(positions_path, sheet_name=0)

            logger.info("Excel files read. Validating columns...")

            required_chart_cols = ['account_name', 'account_chart']
            required_pos_cols = ['account_name', 'normal_account_position', 'description']

            if not all(col in chart_df.columns for col in required_chart_cols):
                missing_cols = [col for col in required_chart_cols if col not in chart_df.columns]
                logger.error(f"CRITICAL ERROR: Missing required columns in '{chart_path}': {missing_cols}")
                messagebox.showerror("Format File Error", f"Kolom berikut hilang di '{chart_path}':\n{missing_cols}")
                return {}
            if not all(col in pos_df.columns for col in required_pos_cols):
                missing_cols = [col for col in required_pos_cols if col not in pos_df.columns]
                logger.error(f"CRITICAL ERROR: Missing required columns in '{positions_path}': {missing_cols}")
                messagebox.showerror("Format File Error", f"Kolom berikut hilang di '{positions_path}':\n{missing_cols}")
                return {}
            logger.info("Columns validated. Processing and merging data...")

            # --- Robust String Processing ---
            # Process chart_df
            chart_df_processed = chart_df[required_chart_cols].copy()
            # 1. Handle potential NaNs then convert to string
            chart_df_processed['account_name'] = chart_df_processed['account_name'].fillna('').astype(str)
            chart_df_processed['account_chart'] = chart_df_processed['account_chart'].fillna('Unknown').astype(str)
            # 2. Apply strip and lower/capitalize using .str accessor
            chart_df_processed['account_name_lower'] = chart_df_processed['account_name'].str.strip().str.lower()
            chart_df_processed['original_name'] = chart_df_processed['account_name'].str.strip()
            chart_df_processed['account_chart'] = chart_df_processed['account_chart'].str.strip().str.capitalize()
            chart_df_processed = chart_df_processed.drop_duplicates(subset=['account_name_lower'], keep='first')

            # Process pos_df
            pos_df_processed = pos_df[required_pos_cols].copy()
            # 1. Handle NaNs then convert to string
            pos_df_processed['account_name'] = pos_df_processed['account_name'].fillna('').astype(str)
            pos_df_processed['normal_account_position'] = pos_df_processed['normal_account_position'].fillna('Unknown').astype(str)
            pos_df_processed['description'] = pos_df_processed['description'].fillna('').astype(str)
            # 2. Apply strip and lower/capitalize using .str accessor
            pos_df_processed['account_name_lower'] = pos_df_processed['account_name'].str.strip().str.lower()
            pos_df_processed['normal_account_position'] = pos_df_processed['normal_account_position'].str.strip().str.capitalize()
            pos_df_processed['description'] = pos_df_processed['description'].str.strip()
            pos_df_processed = pos_df_processed.drop_duplicates(subset=['account_name_lower'], keep='first')
            # --- End Robust String Processing ---

            # Merge dataframes
            merged = pd.merge(
                chart_df_processed[['account_name_lower', 'original_name', 'account_chart']],
                pos_df_processed[['account_name_lower', 'normal_account_position', 'description']],
                on='account_name_lower',
                how='outer'
            )

            # Fill NaNs *after* merge for columns that might not have matched
            merged['original_name'].fillna(merged['account_name_lower'].str.title(), inplace=True)
            merged['account_chart'].fillna('Unknown', inplace=True)
            merged['normal_account_position'].fillna('Unknown', inplace=True)
            merged['description'].fillna('', inplace=True)

            # Remove rows where account_name_lower might be empty after processing
            merged = merged[merged['account_name_lower'] != '']

            # Log potential issues
            only_in_chart = merged[merged['normal_account_position'] == 'Unknown']['original_name'].tolist()
            only_in_pos = merged[merged['account_chart'] == 'Unknown']['original_name'].tolist()
            if only_in_chart: logger.warning(f"Accounts found ONLY in chart file: {only_in_chart}")
            if only_in_pos: logger.warning(f"Accounts found ONLY in positions file: {only_in_pos}")

            # Set index and create dictionary
            account_dict = merged.set_index('account_name_lower').to_dict('index')

            if not account_dict:
                 logger.error("CRITICAL ERROR: No account data loaded after processing. Dictionary is empty.")
                 return {}

            logger.info(f"Successfully loaded and merged data for {len(account_dict)} unique accounts.")
            logger.debug(f"First 5 account keys loaded: {list(account_dict.keys())[:5]}")
            return account_dict

        except FileNotFoundError as fnf_error:
            logger.error(f"CRITICAL ERROR: File not found - {fnf_error}. Ensure '{fnf_error.filename}' exists.", exc_info=True)
            messagebox.showerror("File Error", f"File Excel tidak ditemukan:\n{fnf_error.filename}")
            return {}
        except KeyError as key_error:
             logger.error(f"CRITICAL ERROR: Column '{key_error}' not found. Check headers in Excel files.", exc_info=True)
             messagebox.showerror("Format File Error", f"Kolom '{key_error}' tidak ditemukan.\nCek nama kolom di file Excel.")
             return {}
        except Exception as e:
            logger.exception(f"CRITICAL ERROR: Failed to load account data: {str(e)}")
            messagebox.showerror("Loading Error", f"Gagal memuat data akun: {str(e)}")
            return {}

    def _load_feedback_data(self) -> List[dict]:
        feedback_data = []
        feedback_path = FILES.get('feedback', 'transactions_feedback.json')
        try:
            if os.path.exists(feedback_path):
                with open(feedback_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    logger.info(f"Berhasil memuat {len(data)} feedback dari {feedback_path}")
                    for item in data:
                        try:
                            if self._valid_transaction(item):
                                t = self._parse_transaction(item)
                                t['source'] = 'feedback'  # Tandai sumbernya
                                feedback_data.append(t)
                        except Exception as e:
                            logger.error(f"Error pada feedback: {str(e)} - Item: {item}")
                            continue
                return feedback_data
            else:
                logger.warning(f"File feedback '{feedback_path}' tidak ditemukan. Lanjutkan tanpa data feedback.")
                return []
        except FileNotFoundError:
            logger.error(f"Error: File feedback '{feedback_path}' tidak ditemukan.")
            return []
        except json.JSONDecodeError as json_err:
            logger.error(f"Error decoding JSON di file feedback '{feedback_path}': {json_err}")
            return []
        except Exception as e:
            logger.error(f"Gagal memuat feedback dari '{feedback_path}': {str(e)}")
            return []

    def _init_models(self):
        try:
            descriptions = [t['description'] for t in self.training_data]
            if not any(descriptions):
                raise ValueError("Data training kosong")

            self.vectorizer = TfidfVectorizer(
                max_features=1000,
                ngram_range=(1, 2),
                stop_words='english'
            )

            X = self.vectorizer.fit_transform(descriptions)

            self.nn = NearestNeighbors(n_neighbors=5, metric='cosine')
            self.nn.fit(X)

            logger.info(f"Model berhasil diinisialisasi dengan {len(descriptions)} data training")
        except Exception as e:
            logger.error(f"Model gagal diinisialisasi: {str(e)}")
            raise
    
# In[9]:


# Cell 3.1 - Metode Loading Transactions
def _load_transactions(self) -> List[dict]:
    transactions = []
    try:
        with open(FILES['transactions'], 'r', encoding='utf-8') as f:
            data = json.load(f)
        logger.info(f"Berhasil memuat {len(data)} transaksi")
        for item in data:
            try:
                if self._valid_transaction(item):
                    t = self._parse_transaction(item)
                    t['source'] = 'ml'
                    transactions.append(t)
                else:
                    logger.warning(f"Transaksi tidak valid: {item}")
            except Exception as e:
                logger.error(f"Error pada transaksi: {str(e)}")
                continue
        logger.info(f"Total transaksi valid: {len(transactions)}")
        return transactions
    except Exception as e:
        logger.error(f"Gagal memuat transaksi: {str(e)}")
        return []

def _valid_transaction(self, item: dict) -> bool:
    required = ['description', 'transaction_type', 'entries']
    if not all(k in item for k in required):
        logger.warning(f"Transaksi kurang field wajib: {item}")
        return False
    for entry in item['entries']:
        if 'type' not in entry or 'description' not in entry:
            logger.warning(f"Entri transaksi tidak valid: {entry}")
            return False
    has_debit = any(e['type'] == 'debit' for e in item['entries'])
    has_credit = any(e['type'] == 'credit' for e in item['entries'])
    if not has_debit or not has_credit:
        logger.warning(f"Transaksi tidak memiliki entri debit atau kredit: {item}")
        return False
    return True

def _parse_transaction(self, item: dict) -> dict:
    try:
        debit = next(e for e in item['entries'] if e['type'] == 'debit')
        credit = next(e for e in item['entries'] if e['type'] == 'credit')
        return {
            'description': item['description'].lower().strip(),
            'transaction_type': item['transaction_type'].lower().strip(),
            'debit': debit['description'].strip().lower(),
            'credit': credit['description'].strip().lower()
        }
    except StopIteration:
        raise ValueError("Entri debit atau kredit tidak ditemukan")
    except KeyError as e:
        raise ValueError(f"Field tidak ditemukan: {str(e)}")

# Tambahkan metode ke kelas AccountingAssistant
AccountingAssistant._load_transactions = _load_transactions
AccountingAssistant._valid_transaction = _valid_transaction
AccountingAssistant._parse_transaction = _parse_transaction


# In[10]:


# Cell 3.2 - Metode Feedback dan Model Initialization
def _init_models(self):
    try:
        descriptions = [t['description'] for t in self.training_data]
        if not any(descriptions) or len(descriptions) < 5:  # Perlu beberapa data untuk training classifier
            logger.warning(f"Data training tidak cukup ({len(descriptions)} sampel). Hanya model KNN yang akan diinisialisasi.")
            self.vectorizer = TfidfVectorizer(
                max_features=1000,  # Kurangi fitur jika data sedikit
                ngram_range=(1, 2),
                stop_words='english'
            )
            X = self.vectorizer.fit_transform(descriptions)
            self.nn = NearestNeighbors(n_neighbors=min(3, len(descriptions)), metric='cosine', algorithm='brute')
            self.nn.fit(X)
            self.classifier_available = False  # Tandai classifier tidak tersedia
            return  # Hentikan inisialisasi classifier

        logger.info(f"Memulai inisialisasi model dengan {len(descriptions)} data training...")

        # Vectorizer (seperti sebelumnya, bisa disesuaikan lagi)
        self.vectorizer = TfidfVectorizer(
            max_features=2000,
            ngram_range=(1, 3),
            stop_words='english',
            min_df=2,
            use_idf=True,
            sublinear_tf=True,
            norm='l2'
        )
        X = self.vectorizer.fit_transform(descriptions)
        logger.info(f"Vectorizer TF-IDF berhasil dibuat dengan {X.shape[1]} fitur.")

        # Nearest Neighbors
        self.nn = NearestNeighbors(n_neighbors=min(10, len(descriptions)),
                                   metric='cosine',
                                   algorithm='brute')
        self.nn.fit(X)
        logger.info("Model Nearest Neighbors (KNN) berhasil dilatih.")

        # Target variables
        y_debit = [t['debit'] for t in self.training_data]
        y_credit = [t['credit'] for t in self.training_data]

        self.classifier_available = False
        # Cek apakah data cukup beragam untuk melatih classifier
        if len(set(y_debit)) > 1 and len(set(y_credit)) > 1:
            logger.info("Data cukup beragam, melatih classifier tambahan...")

            # --- IMPROVEMENT 1: RandomForest dengan Class Weight ---
            self.debit_rf_classifier = RandomForestClassifier(n_estimators=100, max_depth=15, random_state=42, class_weight='balanced', n_jobs=-1)
            self.credit_rf_classifier = RandomForestClassifier(n_estimators=100, max_depth=15, random_state=42, class_weight='balanced', n_jobs=-1)
            self.debit_rf_classifier.fit(X, y_debit)
            self.credit_rf_classifier.fit(X, y_credit)
            logger.info("Model RandomForestClassifier berhasil dilatih.")

            # --- IMPROVEMENT 2: Logistic Regression dengan Class Weight ---
            self.debit_lr_classifier = LogisticRegression(random_state=42, class_weight='balanced', solver='liblinear', max_iter=1000, C=1.0)
            self.credit_lr_classifier = LogisticRegression(random_state=42, class_weight='balanced', solver='liblinear', max_iter=1000, C=1.0)
            self.debit_lr_classifier.fit(X, y_debit)
            self.credit_lr_classifier.fit(X, y_credit)
            logger.info("Model LogisticRegression berhasil dilatih.")

            # --- IMPROVEMENT 3 (Optional): Multinomial Naive Bayes ---
            self.debit_nb_classifier = MultinomialNB(alpha=0.1)  # alpha untuk smoothing
            self.credit_nb_classifier = MultinomialNB(alpha=0.1)
            self.debit_nb_classifier.fit(X, y_debit)
            self.credit_nb_classifier.fit(X, y_credit)
            logger.info("Model MultinomialNB berhasil dilatih.")

            self.classifier_available = True  # Tandai classifier tersedia
        else:
            logger.warning("Variasi data debit/kredit tidak cukup untuk melatih classifier tambahan. Hanya KNN yang akan digunakan untuk prediksi detail.")

        logger.info(f"Inisialisasi model selesai. Classifier tersedia: {self.classifier_available}")

    except Exception as e:
        logger.error(f"Model gagal diinisialisasi: {str(e)}")
        # Set flag agar tidak error saat prediksi
        self.vectorizer = None
        self.nn = None
        self.classifier_available = False
        # raise # Re-raise jika ingin menghentikan program jika model gagal


# In[11]:


# Cell 3.3 - NLP Analysis dan Rule-Based Prediction (Revisi _nlp_analyzer)
def _nlp_analyzer(self, description: str) -> str:
    """Analisis NLP yang ditingkatkan untuk menangkap kata kerja aksi dan konteks."""
    try:
        # Terjemahkan dan clean input (jika diperlukan, pastikan berfungsi)
        # Untuk contoh ini, kita asumsikan input sudah dalam bahasa Inggris atau tidak perlu terjemahan
        # translated_desc = translate_input(description) # Komentari jika tidak perlu
        clean_desc = description.lower().strip() # Gunakan deskripsi asli jika tidak ada terjemahan

        doc = nlp_spacy(clean_desc)

        # Ekstrak komponen penting
        important_tokens = []
        action_verbs = []
        nouns_objects = []
        entities = [ent.text for ent in doc.ents if ent.label_ not in ['DATE', 'TIME', 'PERCENT', 'MONEY', 'QUANTITY', 'ORDINAL', 'CARDINAL']] # Abaikan entitas numerik/tanggal

        # Identifikasi kata kerja aksi utama dan objeknya
        for token in doc:
            # Prioritaskan kata kerja utama (root) atau yang terkait langsung dengan objek
            if token.pos_ == 'VERB' and (token.dep_ == 'ROOT' or 'obj' in token.dep_):
                 # Beri bobot lebih tinggi dengan mengulanginya
                 action_verbs.extend([token.lemma_] * 3) # Bobot 3x untuk kata kerja aksi
            # Tangkap kata benda (termasuk proper noun) sebagai objek potensial
            elif token.pos_ in ['NOUN', 'PROPN']:
                 nouns_objects.append(token.lemma_)
            # Tangkap kata sifat yang mungkin memberi konteks penting
            elif token.pos_ == 'ADJ':
                 important_tokens.append(token.lemma_)

        # Gabungkan fitur dengan prioritas: Kata Kerja > Entitas > Objek > Token Penting Lain
        all_features = action_verbs + entities + nouns_objects + important_tokens

        # Hapus duplikat sambil mempertahankan urutan (penting untuk konteks)
        processed_list = []
        seen = set()
        for item in all_features:
            if item not in seen:
                processed_list.append(item)
                seen.add(item)

        processed = " ".join(processed_list)

        # Jika hasil proses kosong (misal hanya angka), kembalikan deskripsi asli
        if not processed.strip():
            processed = clean_desc

        logger.info(f"Input NLP: '{description}' -> Hasil Analisis: '{processed}'")
        return processed

    except Exception as e:
        logger.warning(f"Error dalam NLP analyzer: {str(e)} - fallback ke input original (lowercase)", exc_info=True)
        return description.lower().strip()

# Ganti metode _nlp_analyzer di kelas AccountingAssistant
AccountingAssistant._nlp_analyzer = _nlp_analyzer

# In[12]:


# Cell 3.4 - ML dan Transformer Prediction
def _ml_prediction(self, desc: str) -> dict:
    try:
        # Pemeriksaan awal yang lebih ketat
        if not self.training_data or self.vectorizer is None or self.nn is None:
            logger.error("Model ML tidak terinisialisasi dengan benar atau data training kosong.")
            return {'error': 'Model ML tidak siap untuk prediksi'}

        vec = self.vectorizer.transform([desc])

        # 1. Prediksi KNN (Dasar)
        distances, indices = self.nn.kneighbors(vec)
        if indices.size == 0 or indices[0].size == 0:
            logger.warning(f"KNN tidak menemukan tetangga untuk: '{desc}'")
            return {'error': 'Tidak ditemukan transaksi yang mirip (KNN)'}

        top_knn_predictions = []
        total_knn_weight = 0
        max_neighbors = min(5, len(indices[0]))

        for i in range(max_neighbors):
            idx = indices[0][i]
            if idx >= len(self.training_data): continue

            match = self.training_data[idx]
            raw_confidence = 1 / (distances[0][i] + 1e-6)
            top_knn_predictions.append({
                'debit': match['debit'],
                'credit': match['credit'],
                'confidence': raw_confidence,
                'source_desc': match['description']
            })
            total_knn_weight += raw_confidence

        if not top_knn_predictions:
            logger.warning(f"Tidak ada prediksi KNN yang valid untuk: '{desc}'")
            return {'error': 'Tidak ada prediksi KNN yang valid'}

        knn_debit_votes = {}
        knn_credit_votes = {}
        for pred in top_knn_predictions:
            knn_debit_votes[pred['debit']] = knn_debit_votes.get(pred['debit'], 0) + pred['confidence']
            knn_credit_votes[pred['credit']] = knn_credit_votes.get(pred['credit'], 0) + pred['confidence']

        best_knn_debit = max(knn_debit_votes.items(), key=lambda x: x[1])[0] if knn_debit_votes else None
        best_knn_credit = max(knn_credit_votes.items(), key=lambda x: x[1])[0] if knn_credit_votes else None

        if not best_knn_debit or not best_knn_credit:
             logger.warning(f"KNN tidak dapat menentukan debit/kredit terbaik untuk: '{desc}'")
             return {'error': 'KNN tidak dapat menentukan debit/kredit terbaik'}

        knn_debit_confidence = (knn_debit_votes.get(best_knn_debit, 0) / total_knn_weight) if total_knn_weight > 0 else 0
        knn_credit_confidence = (knn_credit_votes.get(best_knn_credit, 0) / total_knn_weight) if total_knn_weight > 0 else 0
        knn_overall_confidence = min(knn_debit_confidence, knn_credit_confidence)

        # 2. Prediksi Classifier dan Ensemble (HANYA jika classifier tersedia)
        final_debit = best_knn_debit
        final_credit = best_knn_credit
        final_confidence = knn_overall_confidence
        method = 'knn'
        votes_debit = {best_knn_debit: knn_overall_confidence} # Inisialisasi dengan confidence KNN
        votes_credit = {best_knn_credit: knn_overall_confidence}
        classifier_predictions = {}

        if self.classifier_available and \
           self.debit_rf_classifier and self.credit_rf_classifier and \
           self.debit_lr_classifier and self.credit_lr_classifier and \
           self.debit_nb_classifier and self.credit_nb_classifier:

            method = 'ensemble (knn+rf+lr+nb)'
            try:
                classifier_predictions['rf'] = {
                    'debit': self.debit_rf_classifier.predict(vec)[0],
                    'credit': self.credit_rf_classifier.predict(vec)[0]
                }
                classifier_predictions['lr'] = {
                    'debit': self.debit_lr_classifier.predict(vec)[0],
                    'credit': self.credit_lr_classifier.predict(vec)[0]
                }
                classifier_predictions['nb'] = {
                    'debit': self.debit_nb_classifier.predict(vec)[0],
                    'credit': self.credit_nb_classifier.predict(vec)[0]
                }

                # --- IMPROVEMENT: Sesuaikan Bobot Ensemble ---
                # Kurangi sedikit bobot KNN, naikkan bobot classifier
                weights = {'knn': 0.3, 'rf': 0.35, 'lr': 0.25, 'nb': 0.1}

                # Reset votes sebelum menambahkan suara classifier
                votes_debit = {best_knn_debit: weights['knn'] * knn_debit_confidence}
                votes_credit = {best_knn_credit: weights['knn'] * knn_credit_confidence}

                for clf_name, pred in classifier_predictions.items():
                    try:
                         debit_proba = max(getattr(self, f"debit_{clf_name}_classifier").predict_proba(vec)[0])
                         credit_proba = max(getattr(self, f"credit_{clf_name}_classifier").predict_proba(vec)[0])
                         clf_confidence = min(debit_proba, credit_proba)
                    except AttributeError: # Jika model tidak punya predict_proba
                         clf_confidence = 1.0 # Bobot standar jika tidak ada proba

                    votes_debit[pred['debit']] = votes_debit.get(pred['debit'], 0) + weights[clf_name] * clf_confidence
                    votes_credit[pred['credit']] = votes_credit.get(pred['credit'], 0) + weights[clf_name] * clf_confidence

                # Tentukan pemenang voting
                final_debit = max(votes_debit.items(), key=lambda x: x[1])[0] if votes_debit else best_knn_debit
                final_credit = max(votes_credit.items(), key=lambda x: x[1])[0] if votes_credit else best_knn_credit

                # Hitung confidence berdasarkan total bobot pemenang
                total_debit_vote_weight = sum(votes_debit.values())
                total_credit_vote_weight = sum(votes_credit.values())
                ensemble_debit_conf = (votes_debit.get(final_debit, 0) / total_debit_vote_weight) if total_debit_vote_weight > 0 else 0
                ensemble_credit_conf = (votes_credit.get(final_credit, 0) / total_credit_vote_weight) if total_credit_vote_weight > 0 else 0
                final_confidence = min(ensemble_debit_conf, ensemble_credit_conf)

            except Exception as e_clf:
                logger.error(f"Error saat prediksi dengan classifier: {e_clf}", exc_info=True)
                method = 'knn (classifier error)'
                final_confidence = knn_overall_confidence
        else:
            logger.info(f"Classifier tidak tersedia atau tidak terinisialisasi penuh untuk: '{desc}'. Menggunakan hasil KNN saja.")
            method = 'knn (classifier N/A)'

        # Batasi confidence maksimal & cek threshold
        final_confidence = min(float(final_confidence), 0.95)
        if final_confidence < ML_CONFIDENCE_THRESHOLD:
             logger.warning(f"Confidence rendah ({final_confidence:.2f}) untuk '{desc}'. Hasil mungkin tidak akurat.")

        # Siapkan hasil
        result = {
            'debit': final_debit,
            'credit': final_credit,
            'method': method,
            'confidence': final_confidence,
            'details': {
                'knn_prediction': {'debit': best_knn_debit, 'credit': best_knn_credit, 'confidence': knn_overall_confidence},
                'classifier_predictions': classifier_predictions if self.classifier_available else "N/A",
                'debit_votes': votes_debit,
                'credit_votes': votes_credit
            },
            'alternatives': [p for i, p in enumerate(top_knn_predictions) if i > 0 and i < 3]
        }
        logger.info(f"Prediksi ML untuk '{desc}': Debit={final_debit}, Credit={final_credit}, Conf={final_confidence:.2f}, Method={method}")
        return result

    except Exception as e:
        logger.error(f"Error pada prediksi ML: {str(e)}", exc_info=True)
        return {'error': f"Prediksi ML gagal: {str(e)}"}

# Ganti metode _ml_prediction di kelas AccountingAssistant
AccountingAssistant._ml_prediction = _ml_prediction

# In[13]:


# Cell 3.5 - Metode Predict dan Validation

def _rule_based_prediction(self, description: str) -> Optional[dict]:
    """
    Mencoba memprediksi akun debit/kredit berdasarkan aturan keyword dalam Bahasa Inggris.
    Mengembalikan dictionary prediksi jika aturan cocok, None jika tidak.
    Ruleset diperluas untuk mencakup skenario umum dan jurnal penyesuaian.
    """
    # Pastikan description dalam lowercase untuk matching yang case-insensitive
    desc_lower = description.lower().strip()

    # ==============================================================================
    # BAGIAN 1: JURNAL PENYESUAIAN (ADJUSTING ENTRIES)
    # ==============================================================================

    # a. Beban Akrual (Accrued Expenses) - Beban terjadi, kas belum keluar
    if ('accrued' in desc_lower or 'incurred' in desc_lower) and ('expense' in desc_lower or 'payable' in desc_lower):
        if 'salaries' in desc_lower or 'wages' in desc_lower:
            return {'debit': 'Salaries and Wages Expense', 'credit': 'Salaries and Wages Payable', 'method': 'rule_adj_accrued_expense', 'confidence': 1.0}
        if 'interest' in desc_lower:
            return {'debit': 'Interest Expense', 'credit': 'Interest Payable', 'method': 'rule_adj_accrued_expense', 'confidence': 1.0}
        if 'utilities' in desc_lower:
            return {'debit': 'Utilities Expense', 'credit': 'Utilities Payable', 'method': 'rule_adj_accrued_expense', 'confidence': 1.0}

    # b. Pendapatan Akrual (Accrued Revenue) - Pendapatan dihasilkan, kas belum masuk
    if ('accrued' in desc_lower and 'revenue' in desc_lower) or ('services rendered on account' in desc_lower):
        if 'interest' in desc_lower:
            return {'debit': 'Interest Receivable', 'credit': 'Interest Revenue', 'method': 'rule_adj_accrued_revenue', 'confidence': 1.0}
        else: # Untuk pendapatan jasa umum
            return {'debit': 'Accounts Receivable', 'credit': 'Service Revenue', 'method': 'rule_adj_accrued_revenue', 'confidence': 1.0}

    # c. Beban Dibayar Dimuka yang Terpakai (Expired Prepaid Expenses)
    if ('expired' in desc_lower or 'used' in desc_lower or 'consumed' in desc_lower):
        if 'insurance' in desc_lower:
            return {'debit': 'Insurance Expense', 'credit': 'Prepaid Insurance', 'method': 'rule_adj_prepaid_exp', 'confidence': 1.0}
        if 'rent' in desc_lower:
            return {'debit': 'Rent Expense', 'credit': 'Prepaid Rent', 'method': 'rule_adj_prepaid_exp', 'confidence': 1.0}
        if 'supplies' in desc_lower:
            return {'debit': 'Supplies Expense', 'credit': 'Supplies', 'method': 'rule_adj_prepaid_exp', 'confidence': 1.0}

    # d. Pendapatan Diterima Dimuka yang Dihasilkan (Earned Unearned Revenue)
    if ('unearned revenue earned' in desc_lower or 'revenue recognized from advance' in desc_lower):
        return {'debit': 'Unearned Revenue', 'credit': 'Service Revenue', 'method': 'rule_adj_unearned_rev', 'confidence': 1.0}

    # e. Penyusutan (Depreciation)
    if 'depreciation for' in desc_lower or 'record depreciation' in desc_lower:
        if 'equipment' in desc_lower:
            return {'debit': 'Depreciation Expense', 'credit': 'Accumulated Depreciation - Equipment', 'method': 'rule_adj_depreciation', 'confidence': 1.0}
        if 'building' in desc_lower:
            return {'debit': 'Depreciation Expense', 'credit': 'Accumulated Depreciation - Buildings', 'method': 'rule_adj_depreciation', 'confidence': 1.0}
        if 'vehicle' in desc_lower:
            return {'debit': 'Depreciation Expense', 'credit': 'Accumulated Depreciation - Vehicles', 'method': 'rule_adj_depreciation', 'confidence': 1.0}

    # ==============================================================================
    # BAGIAN 2: TRANSAKSI UMUM (PEMBELIAN, PENJUALAN, PEMBAYARAN)
    # ==============================================================================
    
    # --- Aturan Import/Export (Pembelian/Penjualan Barang Dagang) ---
    if ('import' in desc_lower or 'purchase' in desc_lower or 'buy' in desc_lower) and \
       ('product' in desc_lower or 'inventory' in desc_lower or 'goods' in desc_lower or 'merchandise' in desc_lower):
        if 'on account' in desc_lower or 'credit' in desc_lower:
            return {'debit': 'Inventory', 'credit': 'Accounts Payable', 'method': 'rule_import_credit', 'confidence': 1.0}
        elif 'cash' in desc_lower:
            return {'debit': 'Inventory', 'credit': 'Cash', 'method': 'rule_import_cash', 'confidence': 1.0}
        else:
            return {'debit': 'Inventory', 'credit': 'Accounts Payable', 'method': 'rule_import_credit_default', 'confidence': 0.95}

    elif ('export' in desc_lower or 'sell' in desc_lower or 'sale of' in desc_lower) and \
         ('product' in desc_lower or 'inventory' in desc_lower or 'goods' in desc_lower or 'merchandise' in desc_lower):
        if 'on account' in desc_lower or 'credit' in desc_lower or 'billed' in desc_lower:
            return {'debit': 'Accounts Receivable', 'credit': 'Sales Revenue', 'method': 'rule_export_credit', 'confidence': 1.0}
        elif 'cash' in desc_lower:
            return {'debit': 'Cash', 'credit': 'Sales Revenue', 'method': 'rule_export_cash', 'confidence': 1.0}
        else:
            return {'debit': 'Accounts Receivable', 'credit': 'Sales Revenue', 'method': 'rule_export_credit_default', 'confidence': 0.95}

    # --- Transaksi Terkait Persediaan (Retur) ---
    if ('customer return' in desc_lower or 'sales return' in desc_lower):
        if 'cash refund' in desc_lower:
             return {'debit': 'Sales Returns and Allowances', 'credit': 'Cash', 'method': 'rule_sales_return_cash', 'confidence': 1.0}
        else: # Diasumsikan mengurangi piutang
             return {'debit': 'Sales Returns and Allowances', 'credit': 'Accounts Receivable', 'method': 'rule_sales_return_credit', 'confidence': 1.0}
    
    if ('return of goods to supplier' in desc_lower or 'purchase return' in desc_lower):
        if 'cash' in desc_lower:
            return {'debit': 'Cash', 'credit': 'Inventory', 'method': 'rule_purchase_return_cash', 'confidence': 1.0}
        else: # Diasumsikan mengurangi utang
            return {'debit': 'Accounts Payable', 'credit': 'Inventory', 'method': 'rule_purchase_return_credit', 'confidence': 1.0}

    # --- Transaksi Non-Kas & Kredit Lainnya ---
    if ('purchase' in desc_lower or 'bought' in desc_lower) and 'with a note' in desc_lower:
        if 'equipment' in desc_lower:
            return {'debit': 'Equipment', 'credit': 'Notes Payable', 'method': 'rule_asset_on_note', 'confidence': 1.0}
        if 'vehicle' in desc_lower:
            return {'debit': 'Vehicles', 'credit': 'Notes Payable', 'method': 'rule_asset_on_note', 'confidence': 1.0}

    if ('provided services on account' in desc_lower or 'billed a client for services' in desc_lower):
        return {'debit': 'Accounts Receivable', 'credit': 'Service Revenue', 'method': 'rule_service_on_account', 'confidence': 1.0}

    if 'declared a cash dividend' in desc_lower:
        return {'debit': 'Dividends', 'credit': 'Dividends Payable', 'method': 'rule_declare_dividend', 'confidence': 1.0}
        
    # --- Aturan Penerimaan Kas (Sudah ada, bisa di-refine jika perlu) ---
    if 'received cash' in desc_lower or 'collected cash' in desc_lower or 'cash receipt from' in desc_lower:
        if 'from customer' in desc_lower or 'on account' in desc_lower:
            return {'debit': 'Cash', 'credit': 'Accounts Receivable', 'method': 'rule_collect_ar', 'confidence': 1.0}
        if 'issued stock' in desc_lower or 'investment by owner' in desc_lower:
            return {'debit': 'Cash', 'credit': 'Share Capital-Ordinary', 'method': 'rule_issue_stock', 'confidence': 1.0}
        if 'loan proceeds' in desc_lower or 'borrowed' in desc_lower:
            return {'debit': 'Cash', 'credit': 'Notes Payable', 'method': 'rule_get_loan', 'confidence': 1.0}
        if 'interest' in desc_lower: # Tambahan: Penerimaan bunga
            return {'debit': 'Cash', 'credit': 'Interest Revenue', 'method': 'rule_receive_interest', 'confidence': 1.0}
        if 'service' in desc_lower:
            return {'debit': 'Cash', 'credit': 'Service Revenue', 'method': 'rule_cash_service', 'confidence': 0.95}

    # --- Aturan Pembayaran Kas (Sudah ada dan cukup lengkap, dipertahankan) ---
    elif 'paid cash' in desc_lower or 'cash payment for' in desc_lower or 'payment of' in desc_lower:
        # (Blok kode pembayaran kas yang sudah ada dari file asli Anda dapat dipertahankan di sini)
        # ... (contoh beberapa baris)
        if 'rent' in desc_lower:
            return {'debit': 'Rent Expense', 'credit': 'Cash', 'method': 'rule', 'confidence': 1.0}
        elif 'salaries' in desc_lower or 'wages' in desc_lower:
            if 'payable' in desc_lower: # Membayar utang gaji
                return {'debit': 'Salaries and Wages Payable', 'credit': 'Cash', 'method': 'rule', 'confidence': 1.0}
            else: # Membayar beban gaji periode ini
                return {'debit': 'Salaries and Wages Expense', 'credit': 'Cash', 'method': 'rule', 'confidence': 1.0}
        # ... (Lanjutkan dengan aturan pembayaran kas lainnya dari kode asli Anda)
        elif 'dividends' in desc_lower: # Membayar dividen yang sudah diumumkan
            if 'payable' in desc_lower:
                return {'debit': 'Dividends Payable', 'credit': 'Cash', 'method': 'rule', 'confidence': 1.0}
            else: # Langsung bayar tanpa pengumuman (kurang umum di korporasi)
                return {'debit': 'Dividends', 'credit': 'Cash', 'method': 'rule', 'confidence': 0.9}
        # ... dan seterusnya
        
    # Jika tidak ada aturan yang cocok
    return None

def predict(self, description: str) -> dict:
    """Prediksi akun debit dan kredit berdasarkan deskripsi transaksi."""
    if not description:
        return {'error': 'Deskripsi transaksi tidak boleh kosong'}

    # Analisis NLP dulu
    analyzed_desc = self._nlp_analyzer(description)
    logger.info(f"Hasil analisis NLP untuk '{description}': '{analyzed_desc}'")

    # Coba Rule-Based Prediction (Prioritas Tertinggi jika cocok)
    rule_pred = self._rule_based_prediction(analyzed_desc)  # Gunakan hasil NLP
    if rule_pred and 'error' not in rule_pred:
        logger.info(f"Prediksi berbasis ATURAN untuk '{description}'")
        return rule_pred

    logger.info(f"Tidak ada aturan cocok, menggunakan prediksi ML untuk '{description}'")
    # Jika rule tidak cocok atau gagal, gunakan ML Prediction
    ml_pred = self._ml_prediction(analyzed_desc)  # Gunakan hasil NLP

    # Jika tidak ada rule yang cocok, langsung kembalikan hasil ML
    return ml_pred

# Tambahkan metode ke kelas AccountingAssistant
AccountingAssistant.predict = predict
AccountingAssistant._rule_based_prediction = _rule_based_prediction

# In[14]:


# Cell 3.6 - Metode Save Feedback
def _save_feedback(self, description: str, debit: str, credit: str):
    try:
        feedback = {
            'description': description,
            'transaction_type': 'user_feedback',
            'entries': [
                {'description': debit, 'type': 'debit'},
                {'description': credit, 'type': 'credit'}
            ]
        }
        if os.path.exists(FILES['feedback']):
            with open(FILES['feedback'], 'r', encoding='utf-8') as f:
                data = json.load(f)
        else:
            data = []
        data.append(feedback)
        with open(FILES['feedback'], 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4)
        new_feedback = self._parse_transaction(feedback)
        new_feedback['source'] = 'ml'
        boosted_entries = [new_feedback] * FEEDBACK_MULTIPLIER
        self.feedback_data.extend(boosted_entries)
        self.training_data = self.transactions + self.feedback_data
        self._init_models()
        logger.info("Feedback berhasil disimpan dan model diupdate")
    except Exception as e:
        logger.error(f"Gagal menyimpan feedback: {str(e)}")
        raise

# Tambahkan metode ke kelas AccountingAssistant
AccountingAssistant._save_feedback = _save_feedback


# In[15]:


# Cell 4 - AccountingChatbotGUI Kelas Dasar (Revisi __init__ untuk pesan awal)
class AccountingChatbotGUI:
    def __init__(self, assistant: AccountingAssistant, company_name="PT ABCD"):
        # --- Awal Metode __init__ ---
        logger.info("Memulai inisialisasi AccountingChatbotGUI...")  # Log awal
        self.assistant = assistant
        self.company_name = company_name

        self.window = tk.Tk()
        self.window.title(f"{company_name} - AI Accounting Assistant")
        self.window.geometry("900x700")
        self.window.configure(bg='#f0f0f0')

        try:
            self.logo = tk.PhotoImage(file="logo.png")
            self.window.iconphoto(False, self.logo)
        except Exception as e:
            logger.error(f"Gagal memuat logo: {str(e)}")

        # Style Configuration
        self.style = {
            'bg': '#f0f0f0',
            'fg': '#333333',
            'active_bg': '#4CAF50',
            'active_fg': 'white',
            'font': ('Helvetica', 10),
            'title_font': ('Helvetica', 12, 'bold'),
            'header_font': ('Helvetica', 11, 'bold')
        }

        # --- PENTING: Inisialisasi Atribut ---
        self.adjusting_entries = []
        self.financial_data = {
            'income_statement': {},
            'retained_earnings': {},
            'balance_sheet': {}
        }
        self.processed_journal_entries = []  # Inisialisasi di sini
        logger.info("Atribut 'processed_journal_entries' berhasil diinisialisasi sebagai list kosong.")  # Log Konfirmasi

        # --- IMPROVEMENT: Simpan pesan awal ---
        self.initial_journal_messages = [
            ("Sistem", "Selamat datang di AI Accounting Assistant!"),
            ("Sistem", "Silakan masukkan deskripsi transaksi Anda")
        ]
        self.initial_ledger_message = "Pilih file Excel untuk memuat buku besar"
        self.initial_financial_message = "Belum ada data Financial Statement. Silakan load data atau generate dari Buku Besar."
        # --- End Improvement ---

        # Setup Notebook untuk Multiple Tabs
        self.tab_control = ttk.Notebook(self.window)
        self.journal_tab = ttk.Frame(self.tab_control)
        self.ledger_tab = ttk.Frame(self.tab_control)
        self.financial_tab = ttk.Frame(self.tab_control)
        self.tab_control.add(self.journal_tab, text="Prediksi Jurnal")
        self.tab_control.add(self.ledger_tab, text="Buku Besar")
        self.tab_control.add(self.financial_tab, text="Financial Statement")
        self.tab_control.pack(expand=1, fill="both")

        # Initialize semua tab
        logger.info("Memanggil init_journal_tab...")
        self.init_journal_tab()
        logger.info("Memanggil init_ledger_tab...")
        self.init_ledger_tab()
        logger.info("Memanggil init_financial_tab...")
        self.init_financial_tab()

        logger.info("Inisialisasi AccountingChatbotGUI selesai.")
        self.window.mainloop()

# In[16]:


# Cell 4.1 - Tab Prediksi Jurnal
def init_journal_tab(self):
    # Frame utama untuk tab Prediksi Jurnal
    main_frame = tk.Frame(self.journal_tab, bg=self.style['bg'])
    main_frame.pack(fill='both', expand=True, padx=20, pady=20)
    
    # Area Riwayat
    history_label = tk.Label(main_frame, text="Riwayat Interaksi", 
                              font=self.style['title_font'], bg=self.style['bg'])
    history_label.pack(anchor='w', pady=5)
    
    self.history_area = scrolledtext.ScrolledText(main_frame, wrap=tk.WORD,
                                                   width=80, height=20, 
                                                   font=self.style['font'])
    self.history_area.pack(fill='both', expand=True)
    
    # Input Section
    input_frame = tk.Frame(main_frame, bg=self.style['bg'])
    input_frame.pack(fill='x', pady=10)
    
    self.input_var = tk.StringVar()
    self.input_entry = tk.Entry(input_frame, textvariable=self.input_var,
                                 width=60, font=self.style['font'])
    self.input_entry.pack(side='left', padx=5)
    self.input_entry.bind("<Return>", self.handle_enter)  # Binding Enter key
    
    # Tombol Aksi
    btn_frame = tk.Frame(input_frame, bg=self.style['bg'])
    btn_frame.pack(side='right')  # Pastikan frame tombol di kanan

    # Urutkan tombol: Predict, Proses Lanjut, Export, Help, LALU Reset
    self.predict_btn = tk.Button(btn_frame, text="Predict", command=self.handle_prediction,
                                bg='#4CAF50', fg='white', font=self.style['font'])
    self.predict_btn.pack(side='left', padx=2)  # 1. Predict

    self.process_further_btn = tk.Button(btn_frame, text="Proses Lebih Lanjut", 
                                       command=self.open_transaction_selector,
                                       bg='#2196F3', fg='white', font=self.style['font'])
    self.process_further_btn.pack(side='left', padx=2)  # 2. Proses Lanjut

    self.export_to_ledger_btn = tk.Button(btn_frame, text="Export ke Buku Besar", 
                                        command=self.export_to_ledger,
                                        bg='#546E7A', fg='white', font=self.style['font'])
    self.export_to_ledger_btn.pack(side='left', padx=2)  # 3. Export

    self.help_btn = tk.Button(btn_frame, text="Help", command=self.show_help,
                            bg='#9E9E9E', fg='white', font=self.style['font'])
    self.help_btn.pack(side='left', padx=2)  # 4. Help

    # --- Pindahkan Reset ke Akhir ---
    self.reset_journal_btn = tk.Button(btn_frame, text="Reset", command=self.reset_journal_tab,
                                     bg='#FFC107', fg='black', font=self.style['font'])  # Warna kuning
    self.reset_journal_btn.pack(side='left', padx=2)  # 5. Reset (Terakhir)
    # --- End Pindahan ---
    
    # Nominal & Currency
    nominal_frame = tk.Frame(main_frame, bg=self.style['bg'])
    nominal_frame.pack(fill='x', pady=5)
    
    tk.Label(nominal_frame, text="Nominal:", bg=self.style['bg'], font=self.style['font']).pack(side='left')
    self.nominal_var = tk.StringVar()
    self.nominal_entry = tk.Entry(nominal_frame, textvariable=self.nominal_var, width=15, font=self.style['font'])
    self.nominal_entry.pack(side='left', padx=5)
    
    tk.Label(nominal_frame, text="Currency:", bg=self.style['bg'], font=self.style['font']).pack(side='left')
    self.currency_var = tk.StringVar(value="Rupiah")
    self.currency_dropdown = ttk.Combobox(nominal_frame, textvariable=self.currency_var, 
                                        values=["Rupiah", "US Dollar", "Euro"], 
                                        state="readonly", width=10)
    self.currency_dropdown.pack(side='left', padx=5)
    
    # Initial Message
    self.add_history("Sistem", "Selamat datang di AI Accounting Assistant!")
    self.add_history("Sistem", "Silakan masukkan deskripsi transaksi Anda")

# Tambahkan metode ke kelas AccountingChatbotGUI
AccountingChatbotGUI.init_journal_tab = init_journal_tab


# In[17]:


# Cell 4.2 - Tab Buku Besar
def init_ledger_tab(self):
    # Frame utama untuk tab Buku Besar
    main_frame = tk.Frame(self.ledger_tab, bg=self.style['bg'])
    main_frame.pack(fill='both', expand=True, padx=20, pady=20)
    
    # File Input Section
    file_frame = tk.Frame(main_frame, bg=self.style['bg'])
    file_frame.pack(fill='x', pady=10)
    
    tk.Label(file_frame, text="File Excel:", bg=self.style['bg'], font=self.style['font']).pack(side='left')
    self.file_path = tk.StringVar()
    # Beri ruang lebih untuk entry path jika perlu
    tk.Entry(file_frame, textvariable=self.file_path, width=40, font=self.style['font']).pack(side='left', padx=5)

    # Urutkan tombol: Browse, Load, Cek Error, Jurnal Penyesuaian, Fin. Statement, LALU Reset
    self.browse_btn = tk.Button(file_frame, text="Browse", command=self.browse_file,
                               bg='#757575', fg='white', font=self.style['font'])
    self.browse_btn.pack(side='left', padx=2)  # 1. Browse
    
    self.load_btn = tk.Button(file_frame, text="Load", command=self.load_ledger,
                             bg='#4CAF50', fg='white', font=self.style['font'])
    self.load_btn.pack(side='left', padx=2)  # 2. Load
    
    self.check_btn = tk.Button(file_frame, text="Cek Error", command=self.check_ledger_errors_ui,
                              bg='#FF5722', fg='white', font=self.style['font'])
    self.check_btn.pack(side='left', padx=2)  # 3. Cek Error
    
    self.adjust_btn = tk.Button(file_frame, text="Jurnal Penyesuaian", command=self.open_adjusting_entries_ui,
                               bg='#2196F3', fg='white', font=self.style['font'])
    self.adjust_btn.pack(side='left', padx=2)  # 4. Jurnal Penyesuaian
    
    self.export_to_financial_btn = tk.Button(file_frame, text="Fin. Statement",
                                          command=self.export_to_financial_statement,
                                          bg='#FF9800', fg='white', font=self.style['font'])
    self.export_to_financial_btn.pack(side='left', padx=2)  # 5. Fin. Statement
    
    # --- Pindahkan Reset ke Akhir ---
    self.reset_ledger_btn = tk.Button(file_frame, text="Reset", command=self.reset_ledger_tab,
                                     bg='#FFC107', fg='black', font=self.style['font'])  # Warna kuning
    self.reset_ledger_btn.pack(side='left', padx=2)  # 6. Reset (Terakhir)
    # --- End Pindahan ---
    
    # Account Selection Section
    tk.Label(main_frame, text="Pilih Akun:", font=self.style['header_font'], bg=self.style['bg']).pack(anchor='w', pady=(10, 5))
    
    self.account_frame = tk.Frame(main_frame, bg=self.style['bg'])
    self.account_frame.pack(fill='x', pady=5)
    
    # Scrollable frame untuk tombol akun
    self.account_canvas = tk.Canvas(self.account_frame, height=50, bg=self.style['bg'])
    self.account_scrollbar = tk.Scrollbar(self.account_frame, orient="horizontal", command=self.account_canvas.xview)
    self.account_buttons_frame = tk.Frame(self.account_canvas, bg=self.style['bg'])
    
    self.account_buttons_frame.bind("<Configure>", lambda e: self.account_canvas.configure(scrollregion=self.account_canvas.bbox("all")))
    self.account_canvas.create_window((0, 0), window=self.account_buttons_frame, anchor="nw")
    self.account_canvas.configure(xscrollcommand=self.account_scrollbar.set)
    
    self.account_canvas.pack(fill='x', expand=True)
    self.account_scrollbar.pack(fill='x')
    
    # Adjustment Status Label
    self.adjustment_status_frame = tk.Frame(main_frame, bg='#1976D2')  # Ganti background dengan biru
    self.adjustment_status_frame.pack(fill='x', pady=5)
    
    self.adjustment_status_var = tk.StringVar(value="Status: Tidak ada jurnal penyesuaian")
    self.adjustment_status_label = tk.Label(self.adjustment_status_frame, 
                                          textvariable=self.adjustment_status_var,
                                          font=self.style['font'], fg='white', bg='#1976D2')  # Ubah warna teks menjadi putih
    self.adjustment_status_label.pack(side='left', padx=10, pady=5)
    
    self.view_adj_btn = tk.Button(self.adjustment_status_frame, text="Lihat Jurnal Penyesuaian", 
                                 command=self.view_adjusting_entries,
                                 state='disabled',
                                 bg='#0D47A1', fg='white', font=self.style['font'])  # Ganti ungu menjadi biru tua
    self.view_adj_btn.pack(side='right', padx=10, pady=5)
    
    # Ledger Display Section
    tk.Label(main_frame, text="Buku Besar:", font=self.style['header_font'], bg=self.style['bg']).pack(anchor='w', pady=(10, 5))
    
    ledger_frame = tk.Frame(main_frame, bg='white', bd=1, relief='solid')
    ledger_frame.pack(fill='both', expand=True)
    
    # Canvas dan Scrollbar untuk Ledger
    self.ledger_canvas = tk.Canvas(ledger_frame, bg='white')
    self.vsb = tk.Scrollbar(ledger_frame, orient="vertical", command=self.ledger_canvas.yview)
    self.hsb = tk.Scrollbar(ledger_frame, orient="horizontal", command=self.ledger_canvas.xview)
    self.ledger_content_frame = tk.Frame(self.ledger_canvas, bg='white')
    
    self.ledger_content_frame.bind("<Configure>", lambda e: self.ledger_canvas.configure(scrollregion=self.ledger_canvas.bbox("all")))
    self.ledger_canvas.create_window((0, 0), window=self.ledger_content_frame, anchor="nw")
    self.ledger_canvas.configure(yscrollcommand=self.vsb.set, xscrollcommand=self.hsb.set)
    
    self.vsb.pack(side="right", fill="y")
    self.hsb.pack(side="bottom", fill="x")
    self.ledger_canvas.pack(side="left", fill="both", expand=True)
    
    # Default text
    self.guide_label = tk.Label(self.ledger_content_frame, text="Pilih file Excel untuk memuat buku besar",
                              font=self.style['title_font'], bg='white')
    self.guide_label.pack(pady=50)

def settings_menu(self):
    # Buat jendela pengaturan
    settings_window = tk.Toplevel(self.root)
    settings_window.title("Pengaturan Aplikasi")
    settings_window.geometry("400x300")
    settings_window.configure(bg=self.style['bg'])
    
    # Frame untuk pengaturan perusahaan
    company_frame = tk.LabelFrame(settings_window, text="Informasi Perusahaan", 
                                 bg=self.style['bg'], font=('Helvetica', 10, 'bold'))
    company_frame.pack(fill='x', padx=20, pady=20)
    
    # Label dan Entry untuk nama perusahaan
    tk.Label(company_frame, text="Nama Perusahaan:", 
            bg=self.style['bg']).grid(row=0, column=0, sticky='w', padx=10, pady=10)
    
    company_name_var = tk.StringVar(value=self.company_name if hasattr(self, 'company_name') else "PT ABCD")
    company_entry = tk.Entry(company_frame, textvariable=company_name_var, width=30)
    company_entry.grid(row=0, column=1, padx=10, pady=10)
    
    # Tombol simpan
    save_btn = tk.Button(settings_window, text="Simpan", 
                        bg='#2196F3', fg='white',
                        command=lambda: self.save_company_settings(company_name_var.get(), settings_window))
    save_btn.pack(pady=20)

def save_company_settings(self, company_name, window):
    # Simpan nama perusahaan
    self.company_name = company_name
    
    # Perbarui judul aplikasi
    self.root.title(f"{company_name} - Sistem Akuntansi")
    
    # Perbarui tampilan nama perusahaan di seluruh aplikasi
    self.update_company_name_display()
    
    # Tampilkan pesan sukses
    messagebox.showinfo("Sukses", "Pengaturan perusahaan berhasil disimpan!")
    window.destroy()

def update_company_name_display(self):
    # Perbarui tampilan nama perusahaan di layar utama
    if hasattr(self, 'company_label'):
        self.company_label.config(text=self.company_name)
    else:
        # Buat label jika belum ada
        self.company_label = tk.Label(self.header_frame, 
                                     text=self.company_name,
                                     font=('Helvetica', 12, 'bold'),
                                     bg=self.style['bg'])
        self.company_label.pack(side='left', padx=10)
    
    # Perbarui nama perusahaan di laporan dan buku besar
    if hasattr(self, 'ledger_company_label'):
        self.ledger_company_label.config(text=self.company_name)

def create_menu(self):
    # ... kode menu yang sudah ada ...
    settings_menu = tk.Menu(self.menu_bar, tearoff=0)
    settings_menu.add_command(label="Pengaturan Perusahaan", command=self.settings_menu)
    self.menu_bar.add_cascade(label="Pengaturan", menu=settings_menu)

# Tambahkan metode ke kelas AccountingChatbotGUI
AccountingChatbotGUI.init_ledger_tab = init_ledger_tab
AccountingChatbotGUI.settings_menu = settings_menu
AccountingChatbotGUI.save_company_settings = save_company_settings
AccountingChatbotGUI.update_company_name_display = update_company_name_display
AccountingChatbotGUI.create_menu = create_menu


# In[18]:


# Cell 4.3 - Metode Interaksi dan Prediksi
def add_history(self, sender, message, tag=None):
    self.history_area.configure(state='normal')
    self.history_area.insert(tk.END, f"{sender}: {message}\n\n", tag)
    self.history_area.configure(state='disabled')
    self.history_area.see(tk.END)

def handle_enter(self, event=None):
    user_text = self.input_var.get().strip()
    if user_text:
        self.add_history("Anda", user_text)
        self.handle_prediction(user_text)
        self.input_var.set('')

def prompt_manual_entry(self, description, predicted_debit, predicted_credit, nominal, currency_prefix):
    """Membuka jendela untuk entri manual ketika confidence rendah."""

    manual_window = tk.Toplevel(self.window)
    manual_window.title("Entri Manual & Feedback")
    manual_window.geometry("600x450") # Sesuaikan ukuran jika perlu
    manual_window.configure(bg=self.style['bg'])
    manual_window.transient(self.window) # Membuat window ini modal relatif terhadap window utama
    manual_window.grab_set() # Mencegah interaksi dengan window utama

    # Frame Utama
    main_frame = tk.Frame(manual_window, bg=self.style['bg'])
    main_frame.pack(fill='both', expand=True, padx=20, pady=15)

    # Tampilkan Deskripsi Asli
    tk.Label(main_frame, text="Deskripsi Transaksi:", font=self.style['header_font'], bg=self.style['bg'], anchor='w').pack(fill='x', pady=(0, 5))
    desc_label = tk.Label(main_frame, text=description, wraplength=550, justify='left', font=self.style['font'], bg='white', relief='solid', bd=1, padx=5, pady=5)
    desc_label.pack(fill='x', pady=(0, 15))

    # Tampilkan Prediksi Awal (jika ada)
    if predicted_debit != 'N/A':
        tk.Label(main_frame, text="Prediksi Awal (Confidence Rendah):", font=self.style['font'], bg=self.style['bg'], anchor='w').pack(fill='x')
        pred_text = f"Debit: {predicted_debit}\nKredit: {predicted_credit}"
        tk.Label(main_frame, text=pred_text, font=self.style['font'], bg='#FFFFE0', justify='left').pack(fill='x', pady=(0, 15)) # Warna kuning muda

    # --- Input Akun Debit ---
    debit_frame = tk.Frame(main_frame, bg=self.style['bg'])
    debit_frame.pack(fill='x', pady=5)
    tk.Label(debit_frame, text="Akun Debit (Benar):", width=20, anchor='w', bg=self.style['bg'], font=self.style['font']).pack(side='left')

    # Ambil daftar akun dari assistant
    account_list = sorted(self.assistant.account_data.keys()) if hasattr(self.assistant, 'account_data') else []
    if not account_list:
        messagebox.showerror("Error", "Daftar akun tidak tersedia.", parent=manual_window)
        manual_window.destroy()
        return

    debit_account_var = tk.StringVar()
    # Set default ke prediksi awal jika ada dan valid
    if predicted_debit.lower() in account_list:
         debit_account_var.set(predicted_debit)

    debit_combo = ttk.Combobox(debit_frame, textvariable=debit_account_var, values=account_list, width=40, font=self.style['font'], state="readonly")
    debit_combo.pack(side='left', fill='x', expand=True)

    # --- Input Akun Kredit ---
    credit_frame = tk.Frame(main_frame, bg=self.style['bg'])
    credit_frame.pack(fill='x', pady=5)
    tk.Label(credit_frame, text="Akun Kredit (Benar):", width=20, anchor='w', bg=self.style['bg'], font=self.style['font']).pack(side='left')

    credit_account_var = tk.StringVar()
    # Set default ke prediksi awal jika ada dan valid
    if predicted_credit.lower() in account_list:
         credit_account_var.set(predicted_credit)

    credit_combo = ttk.Combobox(credit_frame, textvariable=credit_account_var, values=account_list, width=40, font=self.style['font'], state="readonly")
    credit_combo.pack(side='left', fill='x', expand=True)

    # --- Nominal (Display Only) ---
    nominal_frame = tk.Frame(main_frame, bg=self.style['bg'])
    nominal_frame.pack(fill='x', pady=15)
    tk.Label(nominal_frame, text="Nominal:", width=20, anchor='w', bg=self.style['bg'], font=self.style['font']).pack(side='left')
    tk.Label(nominal_frame, text=f"{currency_prefix}{nominal:,.2f}", width=40, anchor='w', bg='#ECEFF1', font=self.style['font'], relief='solid', bd=1).pack(side='left', fill='x', expand=True)

    # --- Tombol Aksi ---
    button_frame = tk.Frame(main_frame, bg=self.style['bg'])
    button_frame.pack(pady=20)

    def save_manual_feedback():
        selected_debit = debit_account_var.get()
        selected_credit = credit_account_var.get()

        if not selected_debit or not selected_credit:
            messagebox.showerror("Error", "Harap pilih akun Debit dan Kredit.", parent=manual_window)
            return
        if selected_debit == selected_credit:
            messagebox.showerror("Error", "Akun Debit dan Kredit tidak boleh sama.", parent=manual_window)
            return

        try:
            # Panggil metode save_feedback dari assistant
            self.assistant._save_feedback(description, selected_debit, selected_credit)

            # --- IMPROVEMENT: Simpan hasil manual ke list untuk export ---
            self.processed_journal_entries.append({
                'description': description,  # Deskripsi asli
                'debit': selected_debit,     # Akun debit manual
                'credit': selected_credit,   # Akun kredit manual
                'nominal': nominal,          # Nominal dari handle_prediction
                'currency': currency_prefix.replace('.', '').replace('Rp', 'Rupiah').replace('US$', 'US Dollar').replace('€', 'Euro'),  # Konversi balik prefix ke nama currency
                'date': datetime.now()       # Tanggal feedback
            })
            logger.info(f"Menambahkan transaksi manual ke antrian export: {description}")
            # --- End Improvement ---

            # Tampilkan hasil manual di history utama
            manual_debit_display = f"{selected_debit} ({currency_prefix}{nominal:,.2f})"
            manual_credit_display = f"{selected_credit} ({currency_prefix}{nominal:,.2f})"
            feedback_response = (f"✍️ Feedback Tersimpan (Entri Manual):\n"
                                 f"   Debit: {manual_debit_display}\n"
                                 f"   Kredit: {manual_credit_display}")
            self.add_history("Sistem", feedback_response, 'feedback')

            messagebox.showinfo("Sukses", "Feedback berhasil disimpan dan model akan diperbarui.", parent=manual_window)
            manual_window.destroy()  # Tutup window manual

        except Exception as e:
            messagebox.showerror("Error", f"Gagal menyimpan feedback: {str(e)}", parent=manual_window)
            logger.error("Error saat menyimpan feedback manual:", exc_info=True)

    save_button = tk.Button(button_frame, text="Simpan Feedback", command=save_manual_feedback, bg='#4CAF50', fg='white', font=self.style['font'], padx=10)
    save_button.pack(side='left', padx=10)

    cancel_button = tk.Button(button_frame, text="Batal", command=manual_window.destroy, bg='#f44336', fg='white', font=self.style['font'], padx=10)
    cancel_button.pack(side='left', padx=10)
        
def handle_prediction(self, text=None):
    # --- IMPROVEMENT: Pastikan atribut ada sebelum digunakan ---
    if not hasattr(self, 'processed_journal_entries'):
        logger.warning("!!! Atribut 'processed_journal_entries' tidak ditemukan di awal handle_prediction. Menginisialisasi ulang...")
        self.processed_journal_entries = []
    # --- End Improvement ---

    # --- (Kode awal: get text, nominal, currency tetap sama) ---
    if text is None:
        text = self.input_var.get().strip()
        if not text:
            messagebox.showwarning("Perhatian", "Masukkan deskripsi transaksi terlebih dahulu")
            return
        # Pindahkan add_history("Anda", text) ke setelah validasi nominal agar tidak tercatat jika nominal error
        # self.add_history("Anda", text)

    original_description = text  # Simpan deskripsi asli

    nominal_text = self.nominal_var.get().strip()
    try:
        nominal_value = float(nominal_text) if nominal_text else 0
    except ValueError:
        self.add_history("Sistem", "❌ Format nominal tidak valid, harus berupa angka", 'error')
        return

    # Tambahkan input user ke history HANYA setelah validasi nominal berhasil
    self.add_history("Anda", original_description)

    currency_choice = self.currency_var.get()
    if currency_choice == "Rupiah":
        currency_prefix = "Rp."
    elif currency_choice == "US Dollar":
        currency_prefix = "US$"
    else:
        currency_prefix = "€"

    try:
        result = self.assistant.predict(original_description)

        if 'error' in result:
            # ... (Kode penanganan error prediksi tetap sama) ...
            error_message = (f"❌ Error Prediksi:\n{result['error']}\n\n"
                             "💡 Tip:\n• Gunakan deskripsi yang lebih detail\n• Sertakan kata kunci seperti 'beli', 'jual', 'sewa'")
            self.add_history("Sistem", error_message, 'error')
            return

        confidence = result.get('confidence', 0)
        debit_pred = result.get('debit', 'N/A')
        credit_pred = result.get('credit', 'N/A')
        method_pred = result.get('method', 'N/A')

        # Format tampilan nominal
        debit_display = f"{debit_pred} ({currency_prefix}{nominal_value:,.2f})"
        credit_display = f"{credit_pred} ({currency_prefix}{nominal_value:,.2f})"

        CONFIDENCE_THRESHOLD_MANUAL = 0.50

        if confidence < CONFIDENCE_THRESHOLD_MANUAL:
            # ... (Kode untuk confidence rendah dan panggil prompt_manual_entry tetap sama) ...
            warning_message = (f"⚠️ Prediksi Kurang Yakin (Confidence: {confidence*100:.1f}%):\n"
                               f"   Debit: {debit_display}\n"
                               f"   Kredit: {credit_display}\n"
                               f"   Metode: {method_pred}\n\n"
                               "Silakan lakukan entri manual untuk akurasi dan feedback.")
            self.add_history("Sistem", warning_message, 'warning')
            self.prompt_manual_entry(original_description, debit_pred, credit_pred, nominal_value, currency_prefix)

        else:
            # Confidence cukup tinggi
            response = (f"📊 Hasil Prediksi (Confidence: {confidence*100:.1f}%):\n"
                        f"Debit: {debit_display}\n"
                        f"Kredit: {credit_display}\n"
                        f"Metode: {method_pred}")
            self.add_history("Sistem", response, 'prediction')

            # Coba append (sekarang atribut dijamin ada)
            try:
                self.processed_journal_entries.append({
                    'description': original_description,
                    'debit': debit_pred,
                    'credit': credit_pred,
                    'nominal': nominal_value,
                    'currency': self.currency_var.get(),
                    'date': datetime.now()
                })
                logger.info(f"Menambahkan transaksi ke antrian export: {original_description}")
            except Exception as e_append:
                logger.error(f"Error saat append ke processed_journal_entries: {e_append}", exc_info=True)
                self.add_history("Sistem", f"❌ ERROR INTERNAL: Gagal menyimpan hasil prediksi ({type(e_append).__name__}).", 'error')

            self.input_var.set('')
            self.nominal_var.set('')

    except Exception as e:
        # ... (Kode penanganan error utama tetap sama) ...
        error_message = f"❌ Terjadi kesalahan saat prediksi: {str(e)}"
        self.add_history("Sistem", error_message, 'error')
        logger.error("Error di handle_prediction:", exc_info=True)

def browse_file(self):
    filepath = filedialog.askopenfilename(
        title="Pilih File Excel Jurnal",
        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
    )
    if filepath:
        self.file_path.set(filepath)
# Tambahkan metode ke kelas AccountingChatbotGUI
AccountingChatbotGUI.add_history = add_history
AccountingChatbotGUI.handle_enter = handle_enter
AccountingChatbotGUI.handle_prediction = handle_prediction
AccountingChatbotGUI.browse_file = browse_file
AccountingChatbotGUI.prompt_manual_entry = prompt_manual_entry


# In[19]:


# Cell 4.4 - Metode Pengelolaan Buku Besar
def load_ledger(self):
    try:
        file_path = self.file_path.get()
        if not file_path or not os.path.exists(file_path):
            messagebox.showerror("Error", "File tidak ditemukan")
            return
            
        # Baca file Excel
        wb = load_workbook(filename=file_path)
        ws = wb.active
        
        # Proses data jurnal umum
        journal_data = []
        current_description = ""
        
        # Mulai dari baris 6 (setelah header)
        for row in ws.iter_rows(min_row=6, values_only=True):
            if not any(row):  # Skip empty rows
                continue
                
            # Asumsikan format: tanggal, keterangan, ref, akun, debit, kredit
            if len(row) >= 6:
                tanggal, keterangan, ref, akun, debit, kredit = row[:6]
                
                # Jika keterangan tidak kosong, simpan sebagai description baru
                if keterangan:
                    current_description = keterangan
                
                # Jika ada akun, tambahkan ke data
                if akun:
                    journal_data.append({
                        'tanggal': tanggal,
                        'keterangan': current_description,
                        'ref': ref,
                        'akun': akun.strip(),
                        'debit': float(debit) if debit else 0,
                        'kredit': float(kredit) if kredit else 0,
                        'is_adjustment': False  # Mark as original entry
                    })
        
        # Cek error pada data sebelum diproses lebih lanjut
        errors, warnings = self.check_ledger_errors(journal_data)
        
        if errors:
            error_message = "File Excel memiliki error:\n\n" + "\n".join(errors)
            display_window = self.create_error_display("ERROR VALIDASI", error_message, "red")
            return
        
        # Dapatkan daftar akun unik
        accounts = sorted(set(item['akun'] for item in journal_data))
        
        # Buat buku besar berdasarkan akun
        self.ledger_data = {}
        for account in accounts:
            self.ledger_data[account] = [item for item in journal_data if item['akun'] == account]
        
        # Simpan data mentah untuk pengecekan error selanjutnya
        self.raw_journal_data = journal_data
        
        # Reset jurnal penyesuaian saat memuat file baru
        self.adjusting_entries = []
        self.update_adjustment_status()
        
        # Tampilkan tombol akun
        self.display_account_buttons()
        
        # Tampilkan pesan sukses dengan peringatan jika ada
        if warnings:
            warning_message = f"Berhasil memuat {len(accounts)} akun dari jurnal umum.\n\nPeringatan:\n" + "\n".join(warnings)
            display_window = self.create_error_display("PERINGATAN VALIDASI", warning_message, "orange")
        else:
            messagebox.showinfo("Sukses", f"Berhasil memuat {len(accounts)} akun dari jurnal umum")
        
    except Exception as e:
        messagebox.showerror("Error", f"Gagal memuat file Excel: {str(e)}")

def check_ledger_errors_ui(self):
    """Menampilkan UI untuk mengecek error pada data buku besar yang telah dimuat"""
    if not hasattr(self, 'raw_journal_data') or not self.raw_journal_data:
        messagebox.showwarning("Peringatan", "Belum ada data buku besar yang dimuat. Mohon load file Excel terlebih dahulu.")
        return
    
    errors, warnings = self.check_ledger_errors(self.raw_journal_data)
    
    if not errors and not warnings:
        messagebox.showinfo("Validasi", "Tidak ditemukan error atau peringatan pada data buku besar.")
        return
    
    # Buat jendela tampilan error/peringatan
    report_window = tk.Toplevel(self.window)
    report_window.title("Hasil Validasi Buku Besar")
    report_window.geometry("600x400")
    
    # Header
    header_frame = tk.Frame(report_window, bg="#f0f0f0")
    header_frame.pack(fill="x", padx=10, pady=10)
    
    tk.Label(header_frame, text="Hasil Validasi Buku Besar", 
            font=("Helvetica", 14, "bold"), bg="#f0f0f0").pack()
    
    # Area error
    if errors:
        error_frame = tk.LabelFrame(report_window, text="Error", fg="red", font=("Helvetica", 12, "bold"))
        error_frame.pack(fill="x", expand=False, padx=10, pady=5)
        
        error_text = scrolledtext.ScrolledText(error_frame, wrap=tk.WORD, height=6, font=self.style['font'])
        error_text.pack(fill="both", expand=True, padx=5, pady=5)
        error_text.insert(tk.END, "\n".join(errors))
        error_text.configure(state="disabled")
    
    # Area peringatan
    if warnings:
        warning_frame = tk.LabelFrame(report_window, text="Peringatan", fg="orange", font=("Helvetica", 12, "bold"))
        warning_frame.pack(fill="x", expand=False, padx=10, pady=5)
        
        warning_text = scrolledtext.ScrolledText(warning_frame, wrap=tk.WORD, height=6, font=self.style['font'])
        warning_text.pack(fill="both", expand=True, padx=5, pady=5)
        warning_text.insert(tk.END, "\n".join(warnings))
        warning_text.configure(state="disabled")
    
    # Summary
    summary_frame = tk.Frame(report_window, bg="#f0f0f0")
    summary_frame.pack(fill="x", padx=10, pady=10)
    
    summary_text = f"Hasil validasi: {len(errors)} error, {len(warnings)} peringatan"
    tk.Label(summary_frame, text=summary_text, font=self.style['font'], bg="#f0f0f0").pack()
    
    # Close button
    tk.Button(report_window, text="Tutup", command=report_window.destroy, 
             bg="#f44336", fg="white", font=self.style['font']).pack(pady=10)

def check_ledger_errors(self, journal_data):
    """Memeriksa error pada data jurnal buku besar
    
    Returns:
        tuple: (errors, warnings) - list of error messages and warning messages
    """
    errors = []
    warnings = []
    
    # 1. Check if debit and credit are balanced
    total_debit = sum(item['debit'] for item in journal_data)
    total_credit = sum(item['kredit'] for item in journal_data)
    
    if abs(total_debit - total_credit) > 0.01:  # Using 0.01 to handle floating point errors
        errors.append(f"UNBALANCED: Jumlah debit ({total_debit:,.2f}) dan kredit ({total_credit:,.2f}) tidak seimbang. Selisih: {abs(total_debit - total_credit):,.2f}")
    
    # 2. Check if account names exist in chart of accounts
    if hasattr(self.assistant, 'account_data') and self.assistant.account_data:
        existing_accounts = set(acc.lower() for acc in self.assistant.account_data.keys())
        used_accounts = set(item['akun'].lower() for item in journal_data)
        
        unknown_accounts = used_accounts - existing_accounts
        if unknown_accounts:
            warnings.append(f"UNKNOWN ACCOUNTS: {len(unknown_accounts)} akun tidak ditemukan dalam chart of accounts:")
            for account in sorted(unknown_accounts):
                warnings.append(f"- {account}")
    else:
        warnings.append("Tidak dapat memeriksa kesesuaian akun karena chart of accounts tidak tersedia")
    
    return errors, warnings

def create_error_display(self, title, message, color):
    """Membuat jendela khusus untuk menampilkan error/warning dengan format tertentu"""
    display_window = tk.Toplevel(self.window)
    display_window.title(title)
    display_window.geometry("500x300")
    
    # Frame judul
    title_frame = tk.Frame(display_window, bg=color)
    title_frame.pack(fill="x")
    
    tk.Label(title_frame, text=title, font=("Helvetica", 12, "bold"), bg=color, fg="white").pack(pady=5)
    
    # Konten pesan
    content_frame = tk.Frame(display_window)
    content_frame.pack(fill="both", expand=True, padx=10, pady=10)
    
    message_area = scrolledtext.ScrolledText(content_frame, wrap=tk.WORD, font=self.style['font'])
    message_area.pack(fill="both", expand=True)
    message_area.insert(tk.END, message)
    message_area.configure(state="disabled")
    
    # Tombol tutup
    tk.Button(display_window, text="Tutup", command=display_window.destroy, 
             bg="#757575", fg="white", font=self.style['font']).pack(pady=10)
    
    return display_window

# Tambahkan metode ke kelas AccountingChatbotGUI
AccountingChatbotGUI.load_ledger = load_ledger
AccountingChatbotGUI.check_ledger_errors_ui = check_ledger_errors_ui
AccountingChatbotGUI.check_ledger_errors = check_ledger_errors
AccountingChatbotGUI.create_error_display = create_error_display


# In[20]:


# Cell 4.5 - Visualisasi Akun Buku Besar
def display_account_buttons(self):
    # Bersihkan frame tombol
    for widget in self.account_buttons_frame.winfo_children():
        widget.destroy()
    
    # Tambahkan tombol untuk setiap akun dengan command yang benar
    for account in self.ledger_data.keys():
        btn = tk.Button(
            self.account_buttons_frame,
            text=account,
            command=lambda acc=account: self.display_account_ledger(acc)  # Pastikan lambda menggunakan acc
        )
        btn.pack(side='left', padx=5, pady=5)

def determine_account_type(self, account_name):
    account_lower = account_name.lower()
    if any(term in account_lower for term in ['kas', 'piutang', 'persediaan', 'aset', 'asset']):
        return 'asset'
    elif any(term in account_lower for term in ['utang', 'kewajiban', 'liability']):
        return 'liability'
    elif any(term in account_lower for term in ['modal', 'saham', 'equity']):
        return 'equity'
    elif any(term in account_lower for term in ['pendapatan', 'penjualan', 'revenue', 'sales']):
        return 'revenue'
    elif any(term in account_lower for term in ['beban', 'biaya', 'expense']):
        return 'expense'
    else:
        return 'asset'  # Default ke aset jika tidak dikenali

def display_account_ledger(self, account):
    # Hapus konten sebelumnya
    for widget in self.ledger_content_frame.winfo_children():
        widget.destroy()
    
    # Header dengan format IFRS
    header_frame = tk.Frame(self.ledger_content_frame, bg='white')
    header_frame.pack(fill='x', pady=10)
    
    # Judul dengan nama akun dan tanggal
    title_text = f"{account}\nPer {datetime.now().strftime('%d %B %Y')}"
    tk.Label(header_frame, text=title_text, 
             font=('Helvetica', 12, 'bold'), bg='white').pack()

    # Buat frame untuk tabel utama dengan pendekatan grid
    table_frame = tk.Frame(self.ledger_content_frame, bg='white')
    table_frame.pack(fill='both', expand=True, padx=5, pady=5)
    
    # Definisi kolom dengan lebar yang tepat
    columns = ["Date", "Description", "Ref", "Debit", "Credit", "Balance"]
    col_widths = [15, 30, 10, 20, 20, 22]  # Kolom nilai dibuat lebih lebar
    
    # Header tabel dengan garis pembatas
    for col, (text, width) in enumerate(zip(columns, col_widths)):
        header_cell = tk.Label(table_frame, text=text, width=width, 
                 font=('Courier New', 10, 'bold'),  # Font monospace
                 relief='ridge', bd=1, bg='#EEEEEE')
        header_cell.grid(row=0, column=col, sticky='ew')
    
    # Proses data transaksi
    account_entries = self.ledger_data.get(account, []).copy()
    for adj_entry in self.adjusting_entries:
        if adj_entry['akun'] == account:
            account_entries.append(adj_entry)
    
    # Sort by date
    account_entries.sort(key=lambda x: x['tanggal'] if isinstance(x['tanggal'], datetime) else datetime.min)
    
    # Hitung saldo
    account_type = self.determine_account_type(account)
    is_debit_normal = account_type in ['asset', 'expense']
    running_balance = 0
    
    # Tampilkan transaksi
    for i, entry in enumerate(account_entries):
        row_bg = '#F5F5F5' if i % 2 == 0 else 'white'
        if entry.get('is_adjustment', False):
            row_bg = '#E3F2FD'  # Biru muda untuk jurnal penyesuaian
        
        # Format tanggal
        date_str = entry['tanggal'].strftime("%d/%m/%Y") if isinstance(entry['tanggal'], datetime) else ""
        
        # Format deskripsi
        desc_text = entry['keterangan']
        if entry.get('is_adjustment', False):
            desc_text = f"[ADJ] {desc_text}"
        
        # Format nilai dan hitung saldo
        debit_amount = entry.get('debit', 0) or 0
        credit_amount = entry.get('kredit', 0) or 0
        
        # Update saldo
        if is_debit_normal:
            running_balance += debit_amount - credit_amount
        else:
            running_balance += credit_amount - debit_amount
        
        # Format nilai dengan presisi
        debit_str = f"Rp{debit_amount:,.2f}" if debit_amount else ""
        credit_str = f"Rp{credit_amount:,.2f}" if credit_amount else ""
        balance_str = f"Rp{abs(running_balance):,.2f} {'DR' if running_balance >= 0 else 'CR'}"
        
        # Tampilkan dengan wrapping dan alignment yang tepat
        tk.Label(table_frame, text=date_str, width=col_widths[0], 
                 bg=row_bg, anchor='w', font=('Courier New', 9)).grid(
            row=i+1, column=0, sticky='w', padx=1, pady=1)
        
        # Deskripsi dengan wrapping untuk teks yang panjang
        desc_label = tk.Label(table_frame, text=desc_text, width=col_widths[1], 
                             bg=row_bg, anchor='w', font=('Helvetica', 9), 
                             wraplength=200, justify='left')
        desc_label.grid(row=i+1, column=1, sticky='w', padx=1, pady=1)
        
        tk.Label(table_frame, text=entry.get('ref', ''), width=col_widths[2], 
                 bg=row_bg, anchor='center', font=('Courier New', 9)).grid(
            row=i+1, column=2, sticky='ew', padx=1, pady=1)
        
        # Nilai moneter dengan font monospace dan rata kanan
        tk.Label(table_frame, text=debit_str, width=col_widths[3], 
                 bg=row_bg, anchor='e', font=('Courier New', 9)).grid(
            row=i+1, column=3, sticky='e', padx=1, pady=1)
        
        tk.Label(table_frame, text=credit_str, width=col_widths[4], 
                 bg=row_bg, anchor='e', font=('Courier New', 9)).grid(
            row=i+1, column=4, sticky='e', padx=1, pady=1)
        
        tk.Label(table_frame, text=balance_str, width=col_widths[5], 
                 bg=row_bg, anchor='e', font=('Courier New', 9)).grid(
            row=i+1, column=5, sticky='e', padx=1, pady=1)
    
    # Baris total
    total_row = len(account_entries) + 1
    total_debit = sum(entry.get('debit', 0) or 0 for entry in account_entries)
    total_credit = sum(entry.get('kredit', 0) or 0 for entry in account_entries)
    
    # Garis pembatas
    separator = tk.Frame(table_frame, height=2, bg='black')
    separator.grid(row=total_row, column=0, columnspan=6, sticky='ew', pady=3)
    
    # Total dengan format yang rapi
    tk.Label(table_frame, text="TOTAL", font=('Courier New', 10, 'bold'), 
             bg='#EEEEEE', anchor='e').grid(
        row=total_row+1, column=0, columnspan=3, sticky='e', padx=1, pady=1)
    
    tk.Label(table_frame, text=f"Rp{total_debit:,.2f}", width=col_widths[3], 
             bg='#EEEEEE', anchor='e', font=('Courier New', 10, 'bold')).grid(
        row=total_row+1, column=3, sticky='e', padx=1, pady=1)
    
    tk.Label(table_frame, text=f"Rp{total_credit:,.2f}", width=col_widths[4], 
             bg='#EEEEEE', anchor='e', font=('Courier New', 10, 'bold')).grid(
        row=total_row+1, column=4, sticky='e', padx=1, pady=1)
    
    balance_display = f"Rp{abs(running_balance):,.2f} {'DR' if running_balance >= 0 else 'CR'}"
    tk.Label(table_frame, text=balance_display, width=col_widths[5], 
             bg='#EEEEEE', anchor='e', font=('Courier New', 10, 'bold')).grid(
        row=total_row+1, column=5, sticky='e', padx=1, pady=1)
    
    # Update scroll region
    self.ledger_canvas.configure(scrollregion=self.ledger_canvas.bbox("all"))

# Registrasi metode ke kelas
AccountingChatbotGUI.display_account_buttons = display_account_buttons
AccountingChatbotGUI.determine_account_type = determine_account_type
AccountingChatbotGUI.display_account_ledger = display_account_ledger


# In[21]:


# Cell 4.6 - Jurnal Penyesuaian (Bagian 1)
def open_adjusting_entries_ui(self):
    """Membuka antarmuka untuk membuat jurnal penyesuaian"""
    if not hasattr(self, 'ledger_data') or not self.ledger_data:
        messagebox.showwarning("Peringatan", "Belum ada data buku besar yang dimuat. Mohon load file Excel terlebih dahulu.")
        return
    
    # Buat window baru
    adj_window = tk.Toplevel(self.window)
    adj_window.title("Jurnal Penyesuaian")
    adj_window.geometry("800x600")
    adj_window.configure(bg='#f0f0f0')
    
    # Header
    header_frame = tk.Frame(adj_window, bg='#455A64')  # Ganti warna ungu dengan abu-abu gelap
    header_frame.pack(fill='x')
    
    tk.Label(header_frame, text="JURNAL PENYESUAIAN", font=('Helvetica', 14, 'bold'), 
            bg='#455A64', fg='white').pack(pady=10)  # Update bg color
    
    # Main content
    content_frame = tk.Frame(adj_window, bg='#f0f0f0')
    content_frame.pack(fill='both', expand=True, padx=20, pady=20)
    
    # Input Form Section
    form_frame = tk.LabelFrame(content_frame, text="Tambah Jurnal Penyesuaian", 
                              font=self.style['header_font'], bg='#f0f0f0')
    form_frame.pack(fill='x', padx=10, pady=10)
    
    # Date
    date_frame = tk.Frame(form_frame, bg='#f0f0f0')
    date_frame.pack(fill='x', padx=10, pady=5)
    
    tk.Label(date_frame, text="Tanggal:", width=15, anchor='w', 
            bg='#f0f0f0', font=self.style['font']).pack(side='left')
    
    # Setup date picker with default to today
    date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
    date_entry = tk.Entry(date_frame, textvariable=date_var, width=15, font=self.style['font'])
    date_entry.pack(side='left', padx=5)
    
    # Description
    desc_frame = tk.Frame(form_frame, bg='#f0f0f0')
    desc_frame.pack(fill='x', padx=10, pady=5)
    
    tk.Label(desc_frame, text="Keterangan:", width=15, anchor='w', 
            bg='#f0f0f0', font=self.style['font']).pack(side='left')
    
    desc_var = tk.StringVar()
    desc_entry = tk.Entry(desc_frame, textvariable=desc_var, width=50, font=self.style['font'])
    desc_entry.pack(side='left', padx=5, fill='x', expand=True)
    
    # Debit Account
    debit_frame = tk.Frame(form_frame, bg='#f0f0f0')
    debit_frame.pack(fill='x', padx=10, pady=5)
    
    tk.Label(debit_frame, text="Akun Debit:", width=15, anchor='w', 
            bg='#f0f0f0', font=self.style['font']).pack(side='left')
    
    # Get account list
    account_list = sorted(self.ledger_data.keys())
    
    debit_var = tk.StringVar()
    debit_combo = ttk.Combobox(debit_frame, textvariable=debit_var, values=account_list, 
                              width=40, font=self.style['font'])
    debit_combo.pack(side='left', padx=5)
    
    # Credit Account
    credit_frame = tk.Frame(form_frame, bg='#f0f0f0')
    credit_frame.pack(fill='x', padx=10, pady=5)
    
    tk.Label(credit_frame, text="Akun Kredit:", width=15, anchor='w', 
            bg='#f0f0f0', font=self.style['font']).pack(side='left')
    
    credit_var = tk.StringVar()
    credit_combo = ttk.Combobox(credit_frame, textvariable=credit_var, values=account_list, 
                               width=40, font=self.style['font'])
    credit_combo.pack(side='left', padx=5)
    
    # Amount
    amount_frame = tk.Frame(form_frame, bg='#f0f0f0')
    amount_frame.pack(fill='x', padx=10, pady=5)
    
    tk.Label(amount_frame, text="Nominal:", width=15, anchor='w', 
            bg='#f0f0f0', font=self.style['font']).pack(side='left')
    
    amount_var = tk.StringVar()
    amount_entry = tk.Entry(amount_frame, textvariable=amount_var, width=20, font=self.style['font'])
    amount_entry.pack(side='left', padx=5)
    
    # Buttons
    btn_frame = tk.Frame(form_frame, bg='#f0f0f0')
    btn_frame.pack(fill='x', padx=10, pady=10)
    
    def add_adjusting_entry():
        # Validate input
        try:
            date_str = date_var.get().strip()
            description = desc_var.get().strip()
            debit_account = debit_var.get().strip()
            credit_account = credit_var.get().strip()
            amount_str = amount_var.get().strip()
            
            if not date_str or not description or not debit_account or not credit_account or not amount_str:
                messagebox.showwarning("Validasi", "Semua field harus diisi")
                return
            
            try:
                entry_date = datetime.strptime(date_str, "%Y-%m-%d")
            except ValueError:
                messagebox.showwarning("Validasi", "Format tanggal tidak valid. Gunakan YYYY-MM-DD")
                return
            
            try:
                amount = float(amount_str)
                if amount <= 0:
                    raise ValueError("Nominal harus lebih dari 0")
            except ValueError:
                messagebox.showwarning("Validasi", "Nominal tidak valid, harus berupa angka positif")
                return
            
            if debit_account == credit_account:
                messagebox.showwarning("Validasi", "Akun debit dan kredit tidak boleh sama")
                return
            
            # Create the adjusting entries
            debit_entry = {
                'tanggal': entry_date,
                'keterangan': description,
                'ref': 'AJE',
                'akun': debit_account,
                'debit': amount,
                'kredit': 0,
                'is_adjustment': True
            }
            
            credit_entry = {
                'tanggal': entry_date,
                'keterangan': description,
                'ref': 'AJE',
                'akun': credit_account,
                'debit': 0,
                'kredit': amount,
                'is_adjustment': True
            }
            
            # Tambahkan ke adjusting_entries
            self.adjusting_entries.append(debit_entry)
            self.adjusting_entries.append(credit_entry)
            
            # Refresh the display
            self.update_adjustment_status()
            self.refresh_adjusting_entries_list(list_frame)
            
            # Clear input
            desc_var.set('')
            amount_var.set('')
            
            messagebox.showinfo("Sukses", "Jurnal penyesuaian berhasil ditambahkan")
            
        except Exception as e:
            messagebox.showerror("Error", f"Terjadi kesalahan: {str(e)}")
    
    add_btn = tk.Button(btn_frame, text="Tambahkan", command=add_adjusting_entry,
                       bg='#4CAF50', fg='white', font=self.style['font'])
    add_btn.pack(side='right', padx=5)
    
    # Existing Adjusting Entries Display
    list_frame = tk.LabelFrame(content_frame, text="Daftar Jurnal Penyesuaian", 
                              font=self.style['header_font'], bg='#f0f0f0')
    list_frame.pack(fill='both', expand=True, padx=10, pady=10)
    
    # Initialize the list
    self.refresh_adjusting_entries_list(list_frame)
    
    # Button frame
    bottom_frame = tk.Frame(adj_window, bg='#f0f0f0')
    bottom_frame.pack(fill='x', padx=20, pady=10)
    
    def clear_all_adjustments():
        if not self.adjusting_entries:
            messagebox.showinfo("Info", "Tidak ada jurnal penyesuaian untuk dihapus")
            return
        
        result = messagebox.askyesno("Konfirmasi", "Apakah Anda yakin ingin menghapus semua jurnal penyesuaian?")
        if result:
            self.adjusting_entries = []
            self.update_adjustment_status()
            self.refresh_adjusting_entries_list(list_frame)
            messagebox.showinfo("Sukses", "Semua jurnal penyesuaian telah dihapus")
    
    clear_btn = tk.Button(bottom_frame, text="Hapus Semua", command=clear_all_adjustments,
                         bg='#f44336', fg='white', font=self.style['font'])
    clear_btn.pack(side='left', padx=5)
    
    close_btn = tk.Button(bottom_frame, text="Tutup", command=adj_window.destroy,
                         bg='#757575', fg='white', font=self.style['font'])
    close_btn.pack(side='right', padx=5)

# Tambahkan metode ke kelas AccountingChatbotGUI
AccountingChatbotGUI.open_adjusting_entries_ui = open_adjusting_entries_ui


# In[22]:


# Cell 4.7 - Jurnal Penyesuaian (Bagian 2) dan Metode Help

def refresh_adjusting_entries_list(self, parent_frame):
    """Refresh the display of adjusting entries in the UI"""
    # Clear existing content
    for widget in parent_frame.winfo_children():
        widget.destroy()
    
    if not self.adjusting_entries:
        tk.Label(parent_frame, text="Belum ada jurnal penyesuaian", 
                font=self.style['font'], bg='#f0f0f0').pack(pady=20)
        return
    
    # Create headers
    headers_frame = tk.Frame(parent_frame, bg='#e0e0e0')
    headers_frame.pack(fill='x')
    
    headers = ["Tanggal", "Keterangan", "Akun", "Debit", "Kredit", "Aksi"]
    widths = [12, 25, 25, 10, 10, 8]
    
    for i, (header, width) in enumerate(zip(headers, widths)):
        tk.Label(headers_frame, text=header, width=width, 
                bg='#e0e0e0', font=self.style['font']).grid(row=0, column=i, padx=1, pady=2)
    
    # Create scrollable frame for entries
    canvas = tk.Canvas(parent_frame, bg='#f0f0f0')
    scrollbar = tk.Scrollbar(parent_frame, orient="vertical", command=canvas.yview)
    scrollable_frame = tk.Frame(canvas, bg='#f0f0f0')
    
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    # Group adjusting entries by description to show debit and credit together
    entry_groups = {}
    for entry in self.adjusting_entries:
        key = (entry['tanggal'], entry['keterangan'])
        if key not in entry_groups:
            entry_groups[key] = []
        entry_groups[key].append(entry)
    
    # Display entries
    row_idx = 0
    for key, entries in entry_groups.items():
        date, desc = key
        
        # Create a subframe for this entry group
        entry_frame = tk.Frame(scrollable_frame, bg='#f0f0f0')
        entry_frame.pack(fill='x')
        
        # Format date
        date_str = date.strftime("%Y-%m-%d") if isinstance(date, datetime) else str(date)
        
        # Get debit and credit entries
        debit_entries = [e for e in entries if e['debit'] > 0]
        credit_entries = [e for e in entries if e['kredit'] > 0]
        
        # Count how many rows we'll need
        row_count = max(len(debit_entries), len(credit_entries))
        
        for i in range(row_count):
            row_bg = '#f0f0f0' if row_idx % 2 == 0 else '#e0e0e0'
            row_frame = tk.Frame(entry_frame, bg=row_bg)
            row_frame.pack(fill='x')
            
            # Date (only on first row)
            if i == 0:
                tk.Label(row_frame, text=date_str, width=widths[0], 
                        bg=row_bg, font=self.style['font']).grid(row=0, column=0)
            else:
                tk.Label(row_frame, text="", width=widths[0], 
                        bg=row_bg, font=self.style['font']).grid(row=0, column=0)
            
            # Description (only on first row)
            if i == 0:
                tk.Label(row_frame, text=desc, width=widths[1], anchor='w',
                        bg=row_bg, font=self.style['font']).grid(row=0, column=1)
            else:
                tk.Label(row_frame, text="", width=widths[1], 
                        bg=row_bg, font=self.style['font']).grid(row=0, column=1)
            
            # Account
            if i < len(debit_entries):
                tk.Label(row_frame, text=debit_entries[i]['akun'], width=widths[2], anchor='w',
                        bg=row_bg, font=self.style['font']).grid(row=0, column=2)
                tk.Label(row_frame, text=f"{debit_entries[i]['debit']:,.2f}", width=widths[3], anchor='e',
                        bg=row_bg, font=self.style['font']).grid(row=0, column=3)
                tk.Label(row_frame, text="", width=widths[4],
                        bg=row_bg, font=self.style['font']).grid(row=0, column=4)
            elif i < len(credit_entries):
                tk.Label(row_frame, text=" " + credit_entries[i-len(debit_entries)]['akun'], width=widths[2], anchor='w',
                        bg=row_bg, font=self.style['font']).grid(row=0, column=2)
                tk.Label(row_frame, text="", width=widths[3],
                        bg=row_bg, font=self.style['font']).grid(row=0, column=3)
                tk.Label(row_frame, text=f"{credit_entries[i-len(debit_entries)]['kredit']:,.2f}", width=widths[4], anchor='e',
                        bg=row_bg, font=self.style['font']).grid(row=0, column=4)
            
            # Delete button (only on first row)
            if i == 0:
                delete_btn = tk.Button(row_frame, text="Hapus", 
                                     command=lambda k=key: self.delete_adjusting_entry(k, parent_frame),
                                     bg='#f44336', fg='white', font=('Helvetica', 8))
                delete_btn.grid(row=0, column=5)
            else:
                tk.Label(row_frame, text="", width=widths[5], 
                        bg=row_bg, font=self.style['font']).grid(row=0, column=5)
            
            row_idx += 1

def delete_adjusting_entry(self, key, parent_frame):
    """Delete an adjusting entry by its key (date, description)"""
    date, desc = key
    
    result = messagebox.askyesno("Konfirmasi", f"Hapus jurnal penyesuaian '{desc}'?")
    if result:
        # Filter out entries with this key
        self.adjusting_entries = [e for e in self.adjusting_entries 
                                 if not (e['tanggal'] == date and e['keterangan'] == desc)]
        
        # Update status and refresh display
        self.update_adjustment_status()
        self.refresh_adjusting_entries_list(parent_frame)
        messagebox.showinfo("Sukses", "Jurnal penyesuaian berhasil dihapus")

def update_adjustment_status(self):
    """Update the adjustment status label and enable/disable the view button"""
    if self.adjusting_entries:
        count = len(set((e['tanggal'], e['keterangan']) for e in self.adjusting_entries)) 
        self.adjustment_status_var.set(f"Status: {count} jurnal penyesuaian aktif")
        self.view_adj_btn.config(state='normal')
    else:
        self.adjustment_status_var.set("Status: Tidak ada jurnal penyesuaian")
        self.view_adj_btn.config(state='disabled')

def view_adjusting_entries(self):
    """Show a window with all current adjusting entries"""
    if not self.adjusting_entries:
        messagebox.showinfo("Info", "Tidak ada jurnal penyesuaian untuk ditampilkan")
        return
    
    # Create a new window
    view_window = tk.Toplevel(self.window)
    view_window.title("Daftar Jurnal Penyesuaian")
    view_window.geometry("700x400")
    
    # Frame for entries list
    entries_frame = tk.Frame(view_window, bg='white')
    entries_frame.pack(fill='both', expand=True, padx=10, pady=10)
    
    # Call the same method we use in the adjusting entries UI
    self.refresh_adjusting_entries_list(entries_frame)
    
    # Button to close
    tk.Button(view_window, text="Tutup", command=view_window.destroy,
             bg='#757575', fg='white', font=self.style['font']).pack(pady=10)

def show_help(self):
    """Menampilkan dialog bantuan dengan informasi penggunaan aplikasi"""
    help_content = """🆘 Panduan Pengguna

1. Prediksi Jurnal:
    • Klik tombol Predict
    • Masukkan deskripsi transaksi, nominal, dan pilih currency
    • Sistem akan menganalisis deskripsi menggunakan NLP dan ML
    • Hasil prediksi ditampilkan dengan akun debit & kredit
    
2. Proses Lebih Lanjut:
    • Klik tombol Proses Lebih Lanjut untuk memilih transaksi
    • Pilih transaksi yang ingin diproses dengan mencentang kotak
    • Klik Proses untuk melanjutkan atau Kembali untuk kembali ke menu utama
    
3. Buku Besar:
    • Klik tab Buku Besar di bagian atas aplikasi
    • Pilih file Excel hasil ekspor jurnal
    • Klik Cek Error untuk memeriksa validitas file Excel
    • Pilih akun untuk melihat buku besar terkait
    
4. Jurnal Penyesuaian:
    • Klik tombol Jurnal Penyesuaian pada tab Buku Besar
    • Isi form untuk menambahkan jurnal penyesuaian baru
    • Jurnal penyesuaian akan ditampilkan dengan latar belakang abu-abu kebiruan di buku besar
    • Jurnal penyesuaian hanya memengaruhi tampilan, bukan file Excel asli"""
    self.show_info_dialog("Panduan Pengguna", help_content)
 
def show_info_dialog(self, title, message):
    """Menampilkan dialog informasi dengan judul dan pesan tertentu"""
    dialog = tk.Toplevel(self.window)
    dialog.title(title)
    dialog.geometry("400x500")
    text_area = scrolledtext.ScrolledText(dialog, wrap=tk.WORD, font=('Helvetica', 10))
    text_area.pack(fill='both', expand=True, padx=10, pady=10)
    text_area.insert(tk.END, message)
    text_area.configure(state='disabled')
    close_btn = tk.Button(dialog, text="Tutup", command=dialog.destroy)
    close_btn.pack(pady=10)

# Tambahkan metode ke kelas AccountingChatbotGUI
AccountingChatbotGUI.refresh_adjusting_entries_list = refresh_adjusting_entries_list
AccountingChatbotGUI.delete_adjusting_entry = delete_adjusting_entry
AccountingChatbotGUI.update_adjustment_status = update_adjustment_status
AccountingChatbotGUI.view_adjusting_entries = view_adjusting_entries
AccountingChatbotGUI.show_help = show_help
AccountingChatbotGUI.show_info_dialog = show_info_dialog


# In[23]:


# Cell 4.8 - TransactionSelectorWindow dan ConfirmationWindow
def open_transaction_selector(self):
    # Mengumpulkan transaksi dari history
    transactions = []
    text = self.history_area.get("1.0", tk.END)
    lines = text.split("\n\n")
    
    current_transaction = {}
    for line in lines:
        if line.startswith("Anda: "):
            current_transaction = {"description": line[6:].strip()}
        elif line.startswith("Sistem: 📊 Hasil Prediksi:"):
            parts = line.split("\n")
            for part in parts:
                if part.startswith("Debit: "):
                    current_transaction["debit"] = part[7:].strip()
                elif part.startswith("Kredit: "):
                    current_transaction["credit"] = part[8:].strip()
                elif part.startswith("Tingkat Kepercayaan: "):
                    current_transaction["confidence"] = part[21:].strip()
            if "debit" in current_transaction and "credit" in current_transaction:
                transactions.append(current_transaction.copy())
    
    if not transactions:
        messagebox.showinfo("Tidak Ada Transaksi", "Tidak ditemukan transaksi yang dapat diproses.")
        return
    
    TransactionSelectorWindow(self, transactions)

# Tambahkan metode ke kelas AccountingChatbotGUI
AccountingChatbotGUI.open_transaction_selector = open_transaction_selector

# Kelas TransactionSelectorWindow
class TransactionSelectorWindow:
    def __init__(self, parent, transactions):
        self.parent = parent
        self.transactions = transactions
        self.selected_transactions = []
        self.checkboxes = []
        
        self.window = tk.Toplevel(parent.window)
        self.window.title("Pemilihan Transaksi")
        self.window.geometry("800x600")
        self.window.configure(bg='#f0f0f0')
        
        # Header frame
        header_frame = tk.Frame(self.window, bg='#f0f0f0')
        header_frame.pack(fill='x', padx=20, pady=10)
        
        header_label = tk.Label(header_frame, text="Pilih Transaksi untuk Diproses", 
                               font=('Helvetica', 14, 'bold'), bg='#f0f0f0')
        header_label.pack(side='left')
        
        # Transactions frame
        transactions_frame = tk.Frame(self.window, bg='white', bd=1, relief='solid')
        transactions_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Headers for columns
        headers_frame = tk.Frame(transactions_frame, bg='#e0e0e0')
        headers_frame.pack(fill='x')
        
        tk.Label(headers_frame, text="Pilih", width=5, bg='#e0e0e0').pack(side='left')
        tk.Label(headers_frame, text="Deskripsi", width=30, bg='#e0e0e0').pack(side='left')
        tk.Label(headers_frame, text="Debit", width=20, bg='#e0e0e0').pack(side='left')
        tk.Label(headers_frame, text="Kredit", width=20, bg='#e0e0e0').pack(side='left')
        tk.Label(headers_frame, text="Kepercayaan", width=15, bg='#e0e0e0').pack(side='left')
        
        # Transaction list scrollable
        self.canvas = tk.Canvas(transactions_frame, bg='white')
        scrollbar = tk.Scrollbar(transactions_frame, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg='white')
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(
                scrollregion=self.canvas.bbox("all")
            )
        )
        
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)
        
        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Populate transactions
        for i, transaction in enumerate(self.transactions):
            frame = tk.Frame(self.scrollable_frame, bg='white')
            frame.pack(fill='x')
            
            var = tk.BooleanVar()
            checkbox = tk.Checkbutton(frame, variable=var, bg='white')
            checkbox.pack(side='left', padx=5)
            
            tk.Label(frame, text=transaction.get('description', ''), width=30, 
                   anchor='w', bg='white').pack(side='left')
            tk.Label(frame, text=transaction.get('debit', ''), width=20, 
                   anchor='w', bg='white').pack(side='left')
            tk.Label(frame, text=transaction.get('credit', ''), width=20, 
                   anchor='w', bg='white').pack(side='left')
            tk.Label(frame, text=transaction.get('confidence', ''), width=15, 
                   anchor='w', bg='white').pack(side='left')
            
            self.checkboxes.append((var, transaction))
            
            # Alternate row colors
            if i % 2 == 1:
                frame.configure(bg='#f5f5f5')
                for widget in frame.winfo_children():
                    widget.configure(bg='#f5f5f5')
        
        # Buttons frame
        buttons_frame = tk.Frame(self.window, bg='#f0f0f0')
        buttons_frame.pack(fill='x', padx=20, pady=10)
        
        self.back_btn = tk.Button(buttons_frame, text="Kembali", command=self.go_back,
                                bg='#9E9E9E', fg='white', padx=10)
        self.back_btn.pack(side='left', padx=5)
        
        self.help_btn = tk.Button(buttons_frame, text="Bantuan", command=self.show_help,
                                bg='#9E9E9E', fg='white', padx=10)
        self.help_btn.pack(side='left', padx=5)
        
        self.process_btn = tk.Button(buttons_frame, text="Proses", command=self.confirm_process,
                                   bg='#4CAF50', fg='white', padx=10)
        self.process_btn.pack(side='right', padx=5)
    
    def go_back(self):
        self.window.destroy()
    
    def show_help(self):
        help_content = """🆘 Panduan Pemilihan Transaksi

1. Pemilihan Transaksi:
 • Centang kotak di sebelah kiri transaksi yang ingin diproses
 • Anda dapat memilih satu atau lebih transaksi
 
2. Tombol "Kembali":
 • Kembali ke menu utama tanpa memproses transaksi
 
3. Tombol "Proses":
 • Memproses transaksi yang dipilih dan menyimpannya ke Excel
 
4. Tombol "Bantuan":
 • Menampilkan panduan penggunaan pemilihan transaksi ini
"""
        messagebox.showinfo("Panduan Pemilihan Transaksi", help_content)
    
    def confirm_process(self):
        selected = [(i, t) for i, (var, t) in enumerate(self.checkboxes) if var.get()]
        
        if not selected:
            messagebox.showinfo("Tidak Ada Transaksi", "Tidak ada transaksi yang dipilih.")
            return
        
        self.selected_transactions = [t for _, t in selected]
        ConfirmationWindow(self)

# Kelas ConfirmationWindow
class ConfirmationWindow:
    def __init__(self, parent):
        self.parent = parent
        
        self.window = tk.Toplevel(parent.window)
        self.window.title("Konfirmasi")
        self.window.geometry("400x200")
        self.window.configure(bg='#f0f0f0')
        
        # Center the window
        self.window.update_idletasks()
        width = self.window.winfo_width()
        height = self.window.winfo_height()
        x = (self.window.winfo_screenwidth() // 2) - (width // 2)
        y = (self.window.winfo_screenheight() // 2) - (height // 2)
        self.window.geometry(f'{width}x{height}+{x}+{y}')
        
        # Content frame
        content_frame = tk.Frame(self.window, bg='#f0f0f0')
        content_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        message = f"Apakah Anda yakin ingin mengekspor {len(parent.selected_transactions)} transaksi terpilih ke file Excel?"
        message_label = tk.Label(content_frame, text=message, font=('Helvetica', 12), 
                               wraplength=350, bg='#f0f0f0')
        message_label.pack(pady=20)
        
        # Buttons frame
        buttons_frame = tk.Frame(content_frame, bg='#f0f0f0')
        buttons_frame.pack(pady=10)
        
        self.process_btn = tk.Button(buttons_frame, text="Proses", command=self.process_transactions,
                                   bg='#4CAF50', fg='white', padx=20)
        self.process_btn.pack(side='left', padx=10)
        
        self.cancel_btn = tk.Button(buttons_frame, text="Batalkan", command=self.window.destroy,
                                  bg='#f44336', fg='white', padx=20)
        self.cancel_btn.pack(side='left', padx=10)
    
    def process_transactions(self):
        try:
            filename = self.export_to_excel()
            messagebox.showinfo("Ekspor Berhasil", f"Transaksi berhasil diekspor ke file {filename}.")
            self.window.destroy()
            self.parent.window.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"Terjadi kesalahan saat mengekspor: {str(e)}")

    def export_to_excel(self):
        # Membuat file Excel dengan format jurnal umum akuntansi
        current_date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"Jurnal_Umum_{current_date}.xlsx"
        )
        
        if not filename:
            raise ValueError("Ekspor dibatalkan oleh pengguna")
        
        # Persiapkan data untuk jurnal umum
        journal_data = []
        for idx, t in enumerate(self.parent.selected_transactions, 1):
            date = datetime.now().strftime("%Y-%m-%d")
            description = t['description']
            
            # Split debit/credit jika ada nominal
            debit_parts = t['debit'].split(' (')
            debit_account = debit_parts[0]
            
            credit_parts = t['credit'].split(' (')
            credit_account = credit_parts[0]
            
            # Get amount if available
            try:
                if len(debit_parts) > 1:
                    amount_str = debit_parts[1].rstrip(')')
                    if 'Rp.' in amount_str:
                        amount = float(amount_str.replace('Rp.', '').replace(',', ''))
                    elif 'US$' in amount_str:
                        amount = float(amount_str.replace('US$', '').replace(',', ''))
                    else:
                        amount = float(amount_str.replace(',', ''))
                else:
                    amount = 0
            except:
                amount = 0
            
            # Add debit entry
            journal_data.append({
                'Tanggal': date,
                'Keterangan': description if idx == 1 else '',
                'Ref': '',
                'Akun': debit_account,
                'Debit': amount,
                'Kredit': ''
            })
            
            # Add credit entry (indented)
            journal_data.append({
                'Tanggal': '',
                'Keterangan': '',
                'Ref': '',
                'Akun': f" {credit_account}",
                'Debit': '',
                'Kredit': amount
            })
        
        # Create workbook and style it
        wb = Workbook()
        ws = wb.active
        ws.title = "Jurnal Umum"
        
        # Add title
        ws.merge_cells('A1:F1')
        cell = ws['A1']
        cell.value = "JURNAL UMUM"
        cell.font = Font(size=14, bold=True)
        cell.alignment = Alignment(horizontal='center')
        
        # Add company name
        ws.merge_cells('A2:F2')
        cell = ws['A2']
        cell.value = "AI Accounting Assistant"
        cell.font = Font(size=12)
        cell.alignment = Alignment(horizontal='center')
        
        # Add period
        ws.merge_cells('A3:F3')
        cell = ws['A3']
        cell.value = f"Periode: {datetime.now().strftime('%B %Y')}"
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal='center')
        
        # Add headers
        headers = ['Tanggal', 'Keterangan', 'Ref', 'Akun', 'Debit', 'Kredit']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        row = 6
        for data in journal_data:
            for col, (key, value) in enumerate(data.items(), 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                
                # Format currency columns
                if col in [5, 6] and value:  # Debit/Kredit columns
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
            row += 1
        
        # Set column widths
        column_widths = {'A': 12, 'B': 30, 'C': 5, 'D': 25, 'E': 15, 'F': 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Add borders
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        for row in ws.iter_rows(min_row=5, max_row=5+len(journal_data)):
            for cell in row:
                cell.border = thin_border
        
        # Save the workbook
        wb.save(filename)
        return filename


# In[24]:


# Cell 4.9 - Tab Financial Statement dan Integrasi Antar Tab

# Update metode __init__ pada AccountingChatbotGUI untuk menambahkan tab Financial Statement
def updated_init(self, assistant: AccountingAssistant, company_name="PT ABCD"):
    self.assistant = assistant
    self.company_name = company_name  # Simpan nama perusahaan
    
    self.window = tk.Tk()
    self.window.title(f"{company_name} - AI Accounting Assistant")
    self.window.geometry("900x700")
    self.window.configure(bg='#f0f0f0')
    
    try:
        self.logo = tk.PhotoImage(file="logo.png")
        self.window.iconphoto(False, self.logo)
    except Exception as e:
        logger.error(f"Gagal memuat logo: {str(e)}")
    
    # Setup Notebook untuk Multiple Tabs
    self.tab_control = ttk.Notebook(self.window)
    
    # Tab 1: Prediksi Jurnal
    self.journal_tab = ttk.Frame(self.tab_control)
    self.tab_control.add(self.journal_tab, text="Prediksi Jurnal")
    
    # Tab 2: Buku Besar
    self.ledger_tab = ttk.Frame(self.tab_control)
    self.tab_control.add(self.ledger_tab, text="Buku Besar")
    
    # Tab 3: Financial Statement (New)
    self.financial_tab = ttk.Frame(self.tab_control)
    self.tab_control.add(self.financial_tab, text="Financial Statement")
    
    self.tab_control.pack(expand=1, fill="both")
    
    # Style Configuration
    self.style = {
        'bg': '#f0f0f0',
        'fg': '#333333',
        'active_bg': '#4CAF50',
        'active_fg': 'white',
        'font': ('Helvetica', 10),
        'title_font': ('Helvetica', 12, 'bold'),
        'header_font': ('Helvetica', 11, 'bold')
    }
    
    # Variable untuk menyimpan jurnal penyesuaian
    self.adjusting_entries = []
    
    # Variable untuk menyimpan data financial statement
    self.financial_data = {
        'income_statement': {},
        'retained_earnings': {},
        'balance_sheet': {}
    }
    
    # Variable untuk menyimpan data transaksi untuk integrasi antar tab
    self.journal_transactions = []
    
    # Initialize semua tab
    self.init_journal_tab()
    self.init_ledger_tab()
    self.init_financial_tab()
    
    self.window.mainloop()

# Ganti metode __init__ asli dengan yang baru
AccountingChatbotGUI.__init__ = updated_init

# Update metode init_journal_tab untuk menambahkan opsi export ke buku besar
def updated_journal_tab(self):
    # Frame utama untuk tab Prediksi Jurnal
    main_frame = tk.Frame(self.journal_tab, bg=self.style['bg'])
    main_frame.pack(fill='both', expand=True, padx=20, pady=20)
    
    # Area Riwayat
    history_label = tk.Label(main_frame, text="Riwayat Interaksi", 
                              font=self.style['title_font'], bg=self.style['bg'])
    history_label.pack(anchor='w', pady=5)
    
    self.history_area = scrolledtext.ScrolledText(main_frame, wrap=tk.WORD,
                                                   width=80, height=20, 
                                                   font=self.style['font'])
    self.history_area.pack(fill='both', expand=True)
    
    # Input Section
    input_frame = tk.Frame(main_frame, bg=self.style['bg'])
    input_frame.pack(fill='x', pady=10)
    
    self.input_var = tk.StringVar()
    self.input_entry = tk.Entry(input_frame, textvariable=self.input_var,
                                 width=60, font=self.style['font'])
    self.input_entry.pack(side='left', padx=5)
    self.input_entry.bind("<Return>", self.handle_enter)  # Binding Enter key
    
    # Tombol Aksi
    btn_frame = tk.Frame(input_frame, bg=self.style['bg'])
    btn_frame.pack(side='right')
    
    self.predict_btn = tk.Button(btn_frame, text="Predict", command=self.handle_prediction,
                                bg='#4CAF50', fg='white', font=self.style['font'])
    self.predict_btn.pack(side='left', padx=2)
    
    self.process_further_btn = tk.Button(btn_frame, text="Proses Lebih Lanjut", 
                                       command=self.open_transaction_selector,
                                       bg='#2196F3', fg='white', font=self.style['font'])
    self.process_further_btn.pack(side='left', padx=2)
    
    # NEW: Tombol untuk export langsung ke buku besar
    self.export_to_ledger_btn = tk.Button(btn_frame, text="Export ke Buku Besar", 
                                        command=self.export_to_ledger,
                                        bg='#546E7A', fg='white', font=self.style['font'])
    self.export_to_ledger_btn.pack(side='left', padx=2)
    
    self.help_btn = tk.Button(btn_frame, text="Help", command=self.show_help,
                            bg='#9E9E9E', fg='white', font=self.style['font'])
    self.help_btn.pack(side='left', padx=2)
    
    # Nominal & Currency
    nominal_frame = tk.Frame(main_frame, bg=self.style['bg'])
    nominal_frame.pack(fill='x', pady=5)
    
    tk.Label(nominal_frame, text="Nominal:", bg=self.style['bg'], font=self.style['font']).pack(side='left')
    self.nominal_var = tk.StringVar()
    self.nominal_entry = tk.Entry(nominal_frame, textvariable=self.nominal_var, width=15, font=self.style['font'])
    self.nominal_entry.pack(side='left', padx=5)
    
    tk.Label(nominal_frame, text="Currency:", bg=self.style['bg'], font=self.style['font']).pack(side='left')
    self.currency_var = tk.StringVar(value="Rupiah")
    self.currency_dropdown = ttk.Combobox(nominal_frame, textvariable=self.currency_var, 
                                        values=["Rupiah", "US Dollar", "Euro"], 
                                        state="readonly", width=10)
    self.currency_dropdown.pack(side='left', padx=5)
    
    # Initial Message
    self.add_history("Sistem", "Selamat datang di AI Accounting Assistant!")
    self.add_history("Sistem", "Silakan masukkan deskripsi transaksi Anda")

# Ganti metode init_journal_tab asli dengan yang baru
AccountingChatbotGUI.init_journal_tab = updated_journal_tab

# Update metode init_ledger_tab untuk menambahkan opsi export ke financial statement
def updated_ledger_tab(self):
    # Frame utama untuk tab Buku Besar
    main_frame = tk.Frame(self.ledger_tab, bg=self.style['bg'])
    main_frame.pack(fill='both', expand=True, padx=20, pady=20)
    
    # File Input Section
    file_frame = tk.Frame(main_frame, bg=self.style['bg'])
    file_frame.pack(fill='x', pady=10)
    
    tk.Label(file_frame, text="File Excel:", bg=self.style['bg'], font=self.style['font']).pack(side='left')
    self.file_path = tk.StringVar()
    tk.Entry(file_frame, textvariable=self.file_path, width=50, font=self.style['font']).pack(side='left', padx=5)
    
    self.browse_btn = tk.Button(file_frame, text="Browse", command=self.browse_file,
                              bg='#757575', fg='white', font=self.style['font'])
    self.browse_btn.pack(side='left', padx=5)
    
    self.load_btn = tk.Button(file_frame, text="Load", command=self.load_ledger,
                            bg='#4CAF50', fg='white', font=self.style['font'])
    self.load_btn.pack(side='left', padx=5)
    
    # Error Check Button
    self.check_btn = tk.Button(file_frame, text="Cek Error", command=self.check_ledger_errors_ui,
                             bg='#FF5722', fg='white', font=self.style['font'])
    self.check_btn.pack(side='left', padx=5)
    
    # Jurnal Penyesuaian Button
    self.adjust_btn = tk.Button(file_frame, text="Jurnal Penyesuaian", command=self.open_adjusting_entries_ui,
                              bg='#2196F3', fg='white', font=self.style['font'])
    self.adjust_btn.pack(side='left', padx=5)
    
    # NEW: Export to Financial Statement Button
    self.export_to_financial_btn = tk.Button(file_frame, text="Fin. Statement",
                                        command=self.export_to_financial_statement,
                                        bg='#FF9800', fg='white', font=self.style['font'])
    self.export_to_financial_btn.pack(side='left', padx=5)
    
    # Account Selection Section
    tk.Label(main_frame, text="Pilih Akun:", font=self.style['header_font'], bg=self.style['bg']).pack(anchor='w', pady=(10,5))
    
    self.account_frame = tk.Frame(main_frame, bg=self.style['bg'])
    self.account_frame.pack(fill='x', pady=5)
    
    # Scrollable frame untuk tombol akun
    self.account_canvas = tk.Canvas(self.account_frame, height=50, bg=self.style['bg'])
    self.account_scrollbar = tk.Scrollbar(self.account_frame, orient="horizontal", command=self.account_canvas.xview)
    self.account_buttons_frame = tk.Frame(self.account_canvas, bg=self.style['bg'])
    
    self.account_buttons_frame.bind("<Configure>", lambda e: self.account_canvas.configure(scrollregion=self.account_canvas.bbox("all")))
    self.account_canvas.create_window((0, 0), window=self.account_buttons_frame, anchor="nw")
    self.account_canvas.configure(xscrollcommand=self.account_scrollbar.set)
    
    self.account_canvas.pack(fill='x', expand=True)
    self.account_scrollbar.pack(fill='x')
    
    # Adjustment Status Label
    self.adjustment_status_frame = tk.Frame(main_frame, bg=self.style['bg'])
    self.adjustment_status_frame.pack(fill='x', pady=5)
    
    self.adjustment_status_var = tk.StringVar(value="Status: Tidak ada jurnal penyesuaian")
    self.adjustment_status_label = tk.Label(self.adjustment_status_frame, 
                                          textvariable=self.adjustment_status_var,
                                          font=self.style['font'], 
                                          fg='#2196F3',  # Ganti dari #455A64 ke biru cerah
                                          bg=self.style['bg'])
    self.adjustment_status_label.pack(side='left')
    
    self.view_adj_btn = tk.Button(self.adjustment_status_frame, 
                                 text="Lihat Jurnal Penyesuaian",
                                 command=self.view_adjusting_entries,
                                 state='disabled',
                                 bg='#BBDEFB',  # Ganti dari #455A64 ke biru muda
                                 fg='#1565C0',  # Warna teks biru tua
                                 font=self.style['font'])
    self.view_adj_btn.pack(side='right')
    
    # Ledger Display Section
    tk.Label(main_frame, text="Buku Besar:", font=self.style['header_font'], bg=self.style['bg']).pack(anchor='w', pady=(10,5))
    
    ledger_frame = tk.Frame(main_frame, bg='white', bd=1, relief='solid')
    ledger_frame.pack(fill='both', expand=True)
    
    # Canvas dan Scrollbar untuk Ledger
    self.ledger_canvas = tk.Canvas(ledger_frame, bg='white')
    self.vsb = tk.Scrollbar(ledger_frame, orient="vertical", command=self.ledger_canvas.yview)
    self.hsb = tk.Scrollbar(ledger_frame, orient="horizontal", command=self.ledger_canvas.xview)
    self.ledger_content_frame = tk.Frame(self.ledger_canvas, bg='white')
    
    self.ledger_content_frame.bind("<Configure>", lambda e: self.ledger_canvas.configure(scrollregion=self.ledger_canvas.bbox("all")))
    self.ledger_canvas.create_window((0, 0), window=self.ledger_content_frame, anchor="nw")
    self.ledger_canvas.configure(yscrollcommand=self.vsb.set, xscrollcommand=self.hsb.set)
    
    self.vsb.pack(side="right", fill="y")
    self.hsb.pack(side="bottom", fill="x")
    self.ledger_canvas.pack(side="left", fill="both", expand=True)
    
    # Default text
    self.guide_label = tk.Label(self.ledger_content_frame, text="Pilih file Excel untuk memuat buku besar",
                              font=self.style['title_font'], bg='white')
    self.guide_label.pack(pady=50)

# Ganti metode init_ledger_tab asli dengan yang baru
AccountingChatbotGUI.init_ledger_tab = updated_ledger_tab


# In[25]:


# Cell 4.10 - Implementasi Tab Financial Statement (Awal)

def init_financial_tab(self):
    """Inisialisasi tab Financial Statement"""
    # Frame utama untuk tab Financial Statement
    main_frame = tk.Frame(self.financial_tab, bg=self.style['bg'])
    main_frame.pack(fill='both', expand=True, padx=20, pady=20)
    
    # Header
    header_frame = tk.Frame(main_frame, bg=self.style['bg'])
    header_frame.pack(fill='x', pady=10)
    
    tk.Label(header_frame, text="Financial Statement Generator", 
            font=('Helvetica', 14, 'bold'), bg=self.style['bg'], fg='#455A64').pack(side='left')  # Ganti biru muda dengan abu-abu gelap
    
    # File Input Section
    file_frame = tk.Frame(main_frame, bg=self.style['bg'])
    file_frame.pack(fill='x', pady=10)
    
    tk.Label(file_frame, text="File Excel:", bg=self.style['bg'], font=self.style['font']).pack(side='left')
    self.financial_file_path = tk.StringVar()
    tk.Entry(file_frame, textvariable=self.financial_file_path, width=50, font=self.style['font']).pack(side='left', padx=5)
    
    self.financial_browse_btn = tk.Button(file_frame, text="Browse", command=self.browse_financial_file,
                                        bg='#757575', fg='white', font=self.style['font'])
    self.financial_browse_btn.pack(side='left', padx=5)
    
    self.financial_load_btn = tk.Button(file_frame, text="Load", command=self.load_financial_data,
                                      bg='#4CAF50', fg='white', font=self.style['font'])
    self.financial_load_btn.pack(side='left', padx=5)
    
    # Tambahkan tombol Reset
    self.reset_financial_btn = tk.Button(file_frame, text="Reset", command=self.reset_financial_tab,
                                        bg='#FFC107', fg='black', font=self.style['font'])  # Warna kuning
    self.reset_financial_btn.pack(side='left', padx=5)
    
    # Status Label
    self.financial_status_var = tk.StringVar(value="Status: Belum ada data yang dimuat")
    self.financial_status_label = tk.Label(main_frame, textvariable=self.financial_status_var,
                                         font=self.style['font'], fg='#1976D2', bg=self.style['bg'])
    self.financial_status_label.pack(anchor='w', pady=5)
    
    # Notebook untuk sub-tabs dalam Financial Statement
    self.financial_notebook = ttk.Notebook(main_frame)
    self.financial_notebook.pack(fill='both', expand=True, pady=10)
    
    # Sub-tabs
    self.income_statement_tab = ttk.Frame(self.financial_notebook)
    self.retained_earnings_tab = ttk.Frame(self.financial_notebook)
    self.balance_sheet_tab = ttk.Frame(self.financial_notebook)
    
    self.financial_notebook.add(self.income_statement_tab, text="Income Statement")
    self.financial_notebook.add(self.retained_earnings_tab, text="Retained Earnings")
    self.financial_notebook.add(self.balance_sheet_tab, text="Balance Sheet")
    
    # Initialize sub-tabs
    self.init_income_statement_tab()
    self.init_retained_earnings_tab()
    self.init_balance_sheet_tab()

# Tambahkan metode ke kelas AccountingChatbotGUI
AccountingChatbotGUI.init_financial_tab = init_financial_tab


# In[26]:


# Cell 4.11 - Inisialisasi Sub-Tab (Income Statement, Retained Earnings, Balance Sheet)

def init_income_statement_tab(self):
    """Inisialisasi tab Income Statement dengan area scrollable."""
    logger.debug("Initializing Income Statement tab")
    # Frame utama untuk tab
    frame = tk.Frame(self.income_statement_tab, bg='white') # bg white agar konsisten
    frame.pack(fill='both', expand=True, padx=10, pady=10)

    # Header (tombol Generate/Export tetap di luar area scroll)
    header_frame = tk.Frame(frame, bg='white')
    header_frame.pack(fill='x', pady=10)
    tk.Label(header_frame, text="Income Statement",
             font=('Helvetica', 12, 'bold'), bg='white').pack(side='left')
    btn_frame = tk.Frame(header_frame, bg='white')
    btn_frame.pack(side='right')
    self.generate_income_btn = tk.Button(btn_frame, text="Generate",
                                         command=self.generate_income_statement,
                                         bg='#4CAF50', fg='white', font=self.style['font'])
    self.generate_income_btn.pack(side='left', padx=5)
    self.export_income_btn = tk.Button(btn_frame, text="Export to Excel",
                                       command=lambda: self.export_statement_to_excel('income'),
                                       bg='#FF9800', fg='white', font=self.style['font'])
    self.export_income_btn.pack(side='left', padx=5)

    # --- Area Konten Scrollable ---
    content_area_frame = tk.Frame(frame, bg='white') # Frame pembungkus canvas+scrollbar
    content_area_frame.pack(fill='both', expand=True, pady=10)

    canvas = tk.Canvas(content_area_frame, bg='white', highlightthickness=0) # bg white, hapus border canvas
    scrollbar = tk.Scrollbar(content_area_frame, orient="vertical", command=canvas.yview)
    # Frame di dalam canvas yang akan menampung konten laporan
    self.scrollable_income_frame = tk.Frame(canvas, bg='white')

    # Konfigurasi agar frame dalam bisa di-scroll
    self.scrollable_income_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")
        )
    )

    # Masukkan frame ke dalam canvas
    canvas.create_window((0, 0), window=self.scrollable_income_frame, anchor="nw")
    # Link scrollbar ke canvas
    canvas.configure(yscrollcommand=scrollbar.set)

    # Tata letak canvas dan scrollbar
    scrollbar.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)
    # --- Akhir Area Konten Scrollable ---

    # Pesan Default (ditempatkan di dalam frame scrollable)
    self.income_default_label = tk.Label(self.scrollable_income_frame, text="Belum ada data Income Statement. Silakan generate terlebih dahulu.",
                                         font=self.style['font'], bg='white')
    self.income_default_label.pack(pady=50, padx=20) # Tambah padding

def init_retained_earnings_tab(self):
    """Inisialisasi tab Retained Earnings dengan area scrollable."""
    logger.debug("Initializing Retained Earnings tab")
    frame = tk.Frame(self.retained_earnings_tab, bg='white')
    frame.pack(fill='both', expand=True, padx=10, pady=10)

    # Header
    header_frame = tk.Frame(frame, bg='white')
    header_frame.pack(fill='x', pady=10)
    tk.Label(header_frame, text="Retained Earnings Statement",
             font=('Helvetica', 12, 'bold'), bg='white').pack(side='left')
    btn_frame = tk.Frame(header_frame, bg='white')
    btn_frame.pack(side='right')
    self.generate_retained_btn = tk.Button(btn_frame, text="Generate",
                                           command=self.generate_retained_earnings,
                                           bg='#4CAF50', fg='white', font=self.style['font'])
    self.generate_retained_btn.pack(side='left', padx=5)
    self.export_retained_btn = tk.Button(btn_frame, text="Export to Excel",
                                         command=lambda: self.export_statement_to_excel('retained'),
                                         bg='#FF9800', fg='white', font=self.style['font'])
    self.export_retained_btn.pack(side='left', padx=5)

    # --- Area Konten Scrollable ---
    content_area_frame = tk.Frame(frame, bg='white')
    content_area_frame.pack(fill='both', expand=True, pady=10)

    canvas = tk.Canvas(content_area_frame, bg='white', highlightthickness=0)
    scrollbar = tk.Scrollbar(content_area_frame, orient="vertical", command=canvas.yview)
    self.scrollable_retained_frame = tk.Frame(canvas, bg='white') # Frame dalam

    self.scrollable_retained_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )
    canvas.create_window((0, 0), window=self.scrollable_retained_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    scrollbar.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)
    # --- Akhir Area Konten Scrollable ---

    # Pesan Default (di dalam frame scrollable)
    self.retained_default_label = tk.Label(self.scrollable_retained_frame, text="Belum ada data Retained Earnings. Silakan generate terlebih dahulu.",
                                           font=self.style['font'], bg='white')
    self.retained_default_label.pack(pady=50, padx=20)

def init_balance_sheet_tab(self):
    """Inisialisasi tab Balance Sheet dengan area scrollable."""
    logger.debug("Initializing Balance Sheet tab")
    frame = tk.Frame(self.balance_sheet_tab, bg='white')
    frame.pack(fill='both', expand=True, padx=10, pady=10)

    # Header
    header_frame = tk.Frame(frame, bg='white') 
    header_frame.pack(fill='x', pady=10)
    tk.Label(header_frame, text="Statement of Financial Position (Balance Sheet)",
             font=('Helvetica', 12, 'bold'), bg='white').pack(side='left')
    btn_frame = tk.Frame(header_frame, bg='white')
    btn_frame.pack(side='right')
    self.generate_balance_btn = tk.Button(btn_frame, text="Generate",
                                          command=self.generate_balance_sheet,
                                          bg='#4CAF50', fg='white', font=self.style['font'])
    self.generate_balance_btn.pack(side='left', padx=5)
    self.export_balance_btn = tk.Button(btn_frame, text="Export to Excel",
                                        command=lambda: self.export_statement_to_excel('balance'),
                                        bg='#FF9800', fg='white', font=self.style['font'])
    self.export_balance_btn.pack(side='left', padx=5)

    # --- Area Konten Scrollable ---
    content_area_frame = tk.Frame(frame, bg='white')
    content_area_frame.pack(fill='both', expand=True, pady=10)

    canvas = tk.Canvas(content_area_frame, bg='white', highlightthickness=0)
    scrollbar = tk.Scrollbar(content_area_frame, orient="vertical", command=canvas.yview)
    self.scrollable_balance_frame = tk.Frame(canvas, bg='white') # Frame dalam

    self.scrollable_balance_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )
    canvas.create_window((0, 0), window=self.scrollable_balance_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    scrollbar.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)
    # --- Akhir Area Konten Scrollable ---

    # Pesan Default (di dalam frame scrollable)
    self.balance_default_label = tk.Label(self.scrollable_balance_frame, text="Belum ada data Balance Sheet. Silakan generate terlebih dahulu.",
                                          font=self.style['font'], bg='white')
    self.balance_default_label.pack(pady=50, padx=20)

# Tambahkan metode ke kelas AccountingChatbotGUI
AccountingChatbotGUI.init_income_statement_tab = init_income_statement_tab
AccountingChatbotGUI.init_retained_earnings_tab = init_retained_earnings_tab
AccountingChatbotGUI.init_balance_sheet_tab = init_balance_sheet_tab


# In[27]:


# Cell 4.12 - Metode Integrasi dan File Handling untuk Financial Statement

def browse_financial_file(self):
    """Browse untuk file Excel financial data"""
    filepath = filedialog.askopenfilename(
        title="Pilih File Excel Financial Data",
        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
    )
    if filepath:
        self.financial_file_path.set(filepath)

def export_to_ledger(self):
    """Export data dari self.processed_journal_entries langsung ke Buku Besar"""

    # --- IMPROVEMENT: Gunakan list internal, bukan history area ---
    if not self.processed_journal_entries:
        messagebox.showinfo("Tidak Ada Transaksi", "Tidak ditemukan transaksi yang siap diproses untuk Buku Besar.")
        return
    # --- End Improvement ---

    # Buat data jurnal untuk buku besar dalam format yang diharapkan
    journal_data = []
    # --- IMPROVEMENT: Iterasi melalui list internal ---
    for entry_data in self.processed_journal_entries:
        date = entry_data['date']
        description = entry_data['description']
        nominal = entry_data['nominal']
        debit_account = entry_data['debit']
        credit_account = entry_data['credit']

        # Tambahkan entri debit
        journal_data.append({
            'tanggal': date,
            'keterangan': description,
            'ref': 'PJ',  # Prediksi Jurnal / Processed Journal
            'akun': debit_account,
            'debit': nominal,
            'kredit': 0,
            'is_adjustment': False
        })

        # Tambahkan entri kredit
        journal_data.append({
            'tanggal': date,
            'keterangan': description,  # Kosongkan untuk baris kredit di Excel, tapi simpan di data internal
            'ref': 'PJ',
            'akun': credit_account,
            'debit': 0,
            'kredit': nominal,
            'is_adjustment': False
        })
    # --- End Improvement ---

    # Buat buku besar berdasarkan akun
    accounts = sorted(set(item['akun'] for item in journal_data))
    self.ledger_data = {}
    for account in accounts:
        self.ledger_data[account] = [item for item in journal_data if item['akun'] == account]

    # Simpan data mentah untuk pengecekan error
    self.raw_journal_data = journal_data

    # Reset jurnal penyesuaian jika perlu (atau biarkan tergantung logika bisnis)
    # self.adjusting_entries = []
    self.update_adjustment_status()

    # Tampilkan tombol akun
    self.display_account_buttons()

    # Beralih ke tab Buku Besar
    self.tab_control.select(1)  # Index 1 adalah tab Buku Besar

    messagebox.showinfo("Sukses", f"Berhasil memproses {len(self.processed_journal_entries)} transaksi ke Buku Besar.")

    # --- IMPROVEMENT: Kosongkan list setelah diproses ---
    self.processed_journal_entries = []
    logger.info("Antrian transaksi untuk export buku besar telah dikosongkan.")
    # --- End Improvement ---

def export_to_financial_statement(self):
    """Export data dari tab Buku Besar langsung ke Financial Statement"""
    if not hasattr(self, 'ledger_data') or not self.ledger_data:
        messagebox.showwarning("Peringatan", "Belum ada data buku besar yang dimuat.")
        return
    
    # Gunakan data buku besar untuk financial statement
    self.prepare_financial_data_from_ledger()
    
    # Update status
    self.financial_status_var.set(f"Status: Data berhasil dimuat dari Buku Besar ({len(self.ledger_data)} akun)")
    
    # Beralih ke tab Financial Statement
    self.tab_control.select(2)  # Index 2 adalah tab Financial Statement
    
    messagebox.showinfo("Sukses", f"Berhasil mengekspor data Buku Besar ke Financial Statement")

def prepare_financial_data_from_ledger(self):
    """Menyiapkan data finansial dari data buku besar menggunakan Chart of Accounts."""
    logger.info("Attempting to prepare financial data from ledger...")
    if not hasattr(self, 'ledger_data') or not self.ledger_data:
        messagebox.showwarning("Peringatan", "Belum ada data buku besar yang dimuat untuk persiapan data finansial.", parent=self.financial_tab)
        logger.error("Failed to prepare financial data: self.ledger_data is missing or empty.")
        return # Jangan lanjutkan jika tidak ada data ledger

    # --- PERIKSA KETERSEDIAAN ACCOUNT_DATA SEBELUM MEMULAI ---
    if not hasattr(self.assistant, 'account_data') or not self.assistant.account_data:
        messagebox.showerror("Error", "Data Chart of Accounts tidak tersedia. Tidak dapat menyiapkan data finansial.", parent=self.financial_tab)
        logger.error("Failed to prepare financial data: self.assistant.account_data is missing or empty.")
        return # Jangan lanjutkan jika chart of accounts tidak ada
    # --- AKHIR PEMERIKSAAN ---

    # Reset data finansial 
    self.financial_data = {
        'income_statement': {},
        'retained_earnings': {'beginning_balance': {'balance': 0, 'type': 'Equity'}}, # Default RE awal = 0
        'balance_sheet': {}
    }
    logger.debug("Financial data reset before preparation from ledger.")

    processed_accounts_count = 0
    skipped_accounts = []
    # Proses akun-akun dari ledger_data
    for account_name_display, entries in self.ledger_data.items():
        account_name_lower = account_name_display.lower().strip()
        account_info = self.assistant.account_data.get(account_name_lower) # Lookup pakai lowercase

        if not account_info:
            logger.warning(f"Account '{account_name_display}' from ledger not found in Chart of Accounts. Skipping.")
            skipped_accounts.append(account_name_display) 
            continue # Lewati akun jika tidak ada di chart

        # Gunakan account_chart dari assistant.account_data
        account_chart_type = account_info.get('account_chart', 'Unknown').strip().capitalize()
        logger.debug(f"Processing account from ledger: '{account_name_display}', Type: {account_chart_type}")

        # Hitung saldo akhir
        debit_total = sum(entry.get('debit', 0) for entry in entries if entry)
        credit_total = sum(entry.get('kredit', 0) for entry in entries if entry)

        # Tentukan saldo normal
        normal_pos = account_info.get('normal_account_position', 'Unknown').strip().capitalize()
        is_debit_normal = normal_pos == 'Debit'
        if normal_pos == 'Unknown':
             is_debit_normal = account_chart_type in ['Assets', 'Expenses']

        if is_debit_normal:
            balance = debit_total - credit_total
        else:
            balance = credit_total - debit_total

        # Kategorikan akun
        account_data_entry = {'balance': balance, 'type': account_chart_type}

        if account_chart_type == 'Revenues':
            self.financial_data['income_statement'][account_name_lower] = account_data_entry
        elif account_chart_type == 'Expenses':
            self.financial_data['income_statement'][account_name_lower] = account_data_entry
        elif account_chart_type == 'Assets':
            self.financial_data['balance_sheet'][account_name_lower] = account_data_entry
        elif account_chart_type == 'Liability':
            self.financial_data['balance_sheet'][account_name_lower] = account_data_entry
        elif account_chart_type == 'Equity':
            # Saldo dari ledger untuk RE dianggap saldo awal jika belum ada dari TB
            if 'retained earnings' in account_name_lower and 'beginning_balance_from_tb' not in self.financial_data['retained_earnings']:
                 self.financial_data['retained_earnings']['beginning_balance'] = account_data_entry
                 logger.debug(f"Setting beginning RE balance from ledger: {balance} for {account_name_display}")
            self.financial_data['balance_sheet'][account_name_lower] = account_data_entry
        else:
             logger.warning(f"Account '{account_name_display}' has unclassified chart type '{account_chart_type}'.")
             # Optionally add to a separate list if needed

        processed_accounts_count += 1

    if skipped_accounts:
         logger.warning(f"Skipped {len(skipped_accounts)} accounts during financial data prep (not found in chart): {skipped_accounts}")
    logger.info(f"Financial data prepared from {processed_accounts_count} found ledger accounts.")

# Tambahkan metode ke kelas AccountingChatbotGUI
AccountingChatbotGUI.browse_financial_file = browse_financial_file
AccountingChatbotGUI.export_to_ledger = export_to_ledger
AccountingChatbotGUI.export_to_financial_statement = export_to_financial_statement
AccountingChatbotGUI.prepare_financial_data_from_ledger = prepare_financial_data_from_ledger


# In[28]:


# Cell 4.13 - Load Financial Data dari Excel dan Pemrosesan Data

def load_financial_data(self):
    """Load data financial dari file Excel"""
    try:
        file_path = self.financial_file_path.get()
        if not file_path or not os.path.exists(file_path):
            messagebox.showerror("Error", "File tidak ditemukan")
            return
            
        # Baca file Excel
        wb = load_workbook(filename=file_path)
        
        # Reset data finansial
        self.financial_data = {
            'income_statement': {},
            'retained_earnings': {},
            'balance_sheet': {}
        }
        
        # Proses worksheet
        if 'Trial Balance' in wb.sheetnames:
            self.load_trial_balance(wb['Trial Balance'])
        elif 'Journal' in wb.sheetnames:
            self.load_journal_data(wb['Journal'])
        else:
            for ws in wb.worksheets:
                # Coba deteksi tipe worksheet dari kontennya
                if self.is_trial_balance_sheet(ws):
                    self.load_trial_balance(ws)
                    break
                elif self.is_journal_sheet(ws):
                    self.load_journal_data(ws)
                    break
            else:
                messagebox.showwarning("Peringatan", "Format Excel tidak dikenali. Harap gunakan template yang sesuai.")
                return
        
        self.financial_status_var.set(f"Status: Data berhasil dimuat dari Excel ({file_path.split('/')[-1]})")
        messagebox.showinfo("Sukses", "Data Financial Statement berhasil dimuat dari Excel")
        
    except Exception as e:
        messagebox.showerror("Error", f"Gagal memuat file Excel: {str(e)}")

def is_trial_balance_sheet(self, worksheet):
    """Deteksi apakah worksheet adalah Trial Balance"""
    keywords = ['trial balance', 'account', 'debit', 'credit']
    for row in worksheet.iter_rows(max_row=10):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if any(keyword in cell.value.lower() for keyword in keywords):
                    return True
    return False

def is_journal_sheet(self, worksheet):
    """Deteksi apakah worksheet adalah Journal"""
    keywords = ['journal', 'date', 'description', 'debit', 'credit']
    for row in worksheet.iter_rows(max_row=10):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if any(keyword in cell.value.lower() for keyword in keywords):
                    return True
    return False

def load_trial_balance(self, worksheet):
    """Load data dari worksheet Trial Balance menggunakan Chart of Accounts."""
    logger.info(f"Attempting to load data from Trial Balance sheet: {worksheet.title}")

    # --- PERIKSA KETERSEDIAAN ACCOUNT_DATA SEBELUM MEMULAI ---
    if not hasattr(self.assistant, 'account_data') or not self.assistant.account_data:
        logger.error("Failed to load Trial Balance: assistant.account_data is missing or empty.")
        raise ValueError("Data Chart of Accounts tidak tersedia di Assistant.")
    # --- AKHIR PEMERIKSAAN ---

    # (Logika pencarian header tetap sama seperti perbaikan sebelumnya)
    header_row = None
    account_col = debit_col = credit_col = None
    for row_idx, row in enumerate(worksheet.iter_rows(max_row=10), 1):
        row_values = [cell.value for cell in row if cell.value]
        if not row_values: continue
        logger.debug(f"Scanning TB Header Row {row_idx}: {row_values}")
        for col_idx, cell in enumerate(row, 1):
            value = cell.value
            if value and isinstance(value, str):
                value_lower = value.lower().strip()
                if 'account' in value_lower and account_col is None: account_col = col_idx
                elif 'debit' in value_lower and debit_col is None: debit_col = col_idx
                elif 'credit' in value_lower and credit_col is None: credit_col = col_idx
        if account_col and debit_col and credit_col:
            header_row = row_idx
            logger.info(f"Trial Balance header found at row {header_row}. Cols: Acct={account_col}, Dr={debit_col}, Cr={credit_col}")
            break
    if not header_row or not account_col or not debit_col or not credit_col:
         logger.error("Header Trial Balance (Account, Debit, Credit) tidak ditemukan atau tidak lengkap.")
         raise ValueError("Header Trial Balance tidak ditemukan/lengkap (Account, Debit, Credit).")


    # Proses data
    processed_accounts_count = 0
    skipped_accounts = []
    # Pastikan reset data finansial sebelum mengisi
    self.financial_data = {
        'income_statement': {},
        'retained_earnings': {'beginning_balance': {'balance': 0, 'type': 'Equity'}}, # Default RE awal = 0
        'balance_sheet': {}
    }
    logger.debug("Financial data reset before loading Trial Balance.")

    for row_idx, row in enumerate(worksheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
        account_cell = row[account_col - 1]
        account_name_display = account_cell.value
        if not account_name_display or not isinstance(account_name_display, str):
            continue

        account_name_display = account_name_display.strip()
        account_name_lower = account_name_display.lower()

        # Ambil debit/kredit
        debit_val = row[debit_col-1].value
        credit_val = row[credit_col-1].value
        try: debit = float(debit_val) if debit_val is not None else 0.0
        except (ValueError, TypeError): debit = 0.0
        try: credit = float(credit_val) if credit_val is not None else 0.0
        except (ValueError, TypeError): credit = 0.0

        # Cari info akun di Chart of Accounts
        account_info = self.assistant.account_data.get(account_name_lower)
        if not account_info:
            logger.warning(f"Account '{account_name_display}' from TB not found in Chart of Accounts. Skipping.")
            skipped_accounts.append(account_name_display)
            continue

        # Gunakan account_chart untuk klasifikasi
        account_chart_type = account_info.get('account_chart', 'Unknown').strip().capitalize()
        normal_pos = account_info.get('normal_account_position', 'Unknown').strip().capitalize()

        is_debit_normal = normal_pos == 'Debit'
        if normal_pos == 'Unknown':
             is_debit_normal = account_chart_type in ['Assets', 'Expenses']

        # Hitung saldo dari TB
        if is_debit_normal:
            balance = debit - credit
        else:
            balance = credit - debit

        # Kategorikan akun
        account_data_entry = {'balance': balance, 'type': account_chart_type}

        if account_chart_type == 'Revenues':
            self.financial_data['income_statement'][account_name_lower] = account_data_entry
        elif account_chart_type == 'Expenses':
             self.financial_data['income_statement'][account_name_lower] = account_data_entry
        elif account_chart_type == 'Assets':
            self.financial_data['balance_sheet'][account_name_lower] = account_data_entry
        elif account_chart_type == 'Liability':
            self.financial_data['balance_sheet'][account_name_lower] = account_data_entry
        elif account_chart_type == 'Equity':
            if 'retained earnings' in account_name_lower:
                 # Saldo RE dari TB adalah saldo awal
                 self.financial_data['retained_earnings']['beginning_balance'] = account_data_entry
                 # Tandai sumbernya jika perlu debugging
                 self.financial_data['retained_earnings']['beginning_balance_from_tb'] = True
                 logger.debug(f"Setting beginning RE balance from TB: {balance} for {account_name_display}")
            self.financial_data['balance_sheet'][account_name_lower] = account_data_entry
        else:
             logger.warning(f"Account '{account_name_display}' has unclassified chart type '{account_chart_type}'. Skipping.")
             skipped_accounts.append(account_name_display)

        processed_accounts_count += 1

    if skipped_accounts:
        logger.warning(f"Skipped {len(skipped_accounts)} accounts during TB load (not found in chart): {skipped_accounts}")
    logger.info(f"Successfully processed {processed_accounts_count} accounts from Trial Balance sheet.")

def load_journal_data(self, worksheet):
    """Load data dari worksheet Journal dan hitung saldo akhir menggunakan Chart of Accounts."""
    logger.info(f"Attempting to load data from Journal sheet: {worksheet.title}")

    # --- PERIKSA KETERSEDIAAN ACCOUNT_DATA SEBELUM MEMULAI ---
    if not hasattr(self.assistant, 'account_data') or not self.assistant.account_data:
        logger.error("Failed to load Journal data: assistant.account_data is missing or empty.")
        raise ValueError("Data Chart of Accounts tidak tersedia di Assistant.")
    # --- AKHIR PEMERIKSAAN ---

    # (Logika pencarian header tetap sama seperti perbaikan sebelumnya)
    header_row = date_col = desc_col = account_col = debit_col = credit_col = None
    for row_idx, row in enumerate(worksheet.iter_rows(max_row=10), 1):
        row_values = [cell.value for cell in row if cell.value]
        if not row_values: continue
        logger.debug(f"Scanning Journal Header Row {row_idx}: {row_values}")
        for col_idx, cell in enumerate(row, 1):
            value = cell.value
            if value and isinstance(value, str):
                value_lower = value.lower().strip()
                if 'date' in value_lower and date_col is None: date_col = col_idx
                if ('description' in value_lower or 'keterangan' in value_lower) and desc_col is None: desc_col = col_idx
                if ('account' in value_lower or 'akun' in value_lower) and account_col is None: account_col = col_idx
                if 'debit' in value_lower and debit_col is None: debit_col = col_idx
                if ('credit' in value_lower or 'kredit' in value_lower) and credit_col is None: credit_col = col_idx
        if account_col and debit_col and credit_col:
            header_row = row_idx
            logger.info(f"Journal header found at row {header_row}. Cols: Acct={account_col}, Dr={debit_col}, Cr={credit_col}")
            break
    if not header_row or not account_col or not debit_col or not credit_col:
         logger.error("Header Journal (Account, Debit, Credit) tidak ditemukan atau tidak lengkap.")
         raise ValueError("Header Journal tidak ditemukan/lengkap (Account, Debit, Credit).")

    # Struktur sementara untuk akumulasi saldo
    account_balances = {}

    # Proses data jurnal
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
        account_cell = row[account_col - 1]
        account_name_display = account_cell.value
        if not account_name_display or not isinstance(account_name_display, str):
            continue

        account_name_display = account_name_display.strip()
        account_name_lower = account_name_display.lower()

        debit_val = row[debit_col-1].value
        credit_val = row[credit_col-1].value
        try: debit = float(debit_val) if debit_val is not None else 0.0
        except (ValueError, TypeError): debit = 0.0
        try: credit = float(credit_val) if credit_val is not None else 0.0
        except (ValueError, TypeError): credit = 0.0

        if account_name_lower not in account_balances:
            account_balances[account_name_lower] = {'debit': 0.0, 'credit': 0.0, 'display_name': account_name_display}
        account_balances[account_name_lower]['debit'] += debit
        account_balances[account_name_lower]['credit'] += credit

    # Hitung saldo akhir dan klasifikasikan
    processed_accounts_count = 0
    skipped_accounts = []
    # Pastikan reset data finansial sebelum mengisi
    self.financial_data = {
        'income_statement': {},
        'retained_earnings': {'beginning_balance': {'balance': 0, 'type': 'Equity'}}, # Default RE awal = 0
        'balance_sheet': {}
    }
    logger.debug("Financial data reset before processing Journal balances.")

    # Proses saldo tiap akun dengan normalisasi nama
    for account_name_lower, totals in account_balances.items():
        # ===== Normalisasi / Mapping Nama Akun =====
        normalized_name = account_name_lower

        # 1. Contoh mapping sederhana untuk kasus plural vs singular:
        manual_map = {
            'account payable':       'accounts payable',
            'account receivable':    'accounts receivables',
            # tambahkan entri lain sesuai kebutuhan
        }
        if account_name_lower in manual_map:
            normalized_name = manual_map[account_name_lower]

        # 2. Jika masih tidak ketemu, coba cari pasangan terdekat di chart:
        if normalized_name not in self.assistant.account_data:
            kandidat = difflib.get_close_matches(normalized_name,
                                                 self.assistant.account_data.keys(),
                                                 n=1, cutoff=0.8)
            if kandidat:
                normalized_name = kandidat[0]

        account_info = self.assistant.account_data.get(normalized_name)
        if not account_info:
            logger.warning(f"Account '{totals['display_name']}' from Journal not found in Chart of Accounts even after normalization. Skipping.")
            skipped_accounts.append(totals['display_name'])
            continue
        # ===========================================

        # Gunakan normalized_name untuk klasifikasi
        account_chart_type = account_info.get('account_chart', 'Unknown').strip().capitalize()
        normal_pos = account_info.get('normal_account_position', 'Unknown').strip().capitalize()

        is_debit_normal = normal_pos == 'Debit'
        if normal_pos == 'Unknown':
            is_debit_normal = account_chart_type in ['Assets', 'Expenses']

        if is_debit_normal:
            balance = totals['debit'] - totals['credit']
        else:
            balance = totals['credit'] - totals['debit']

        account_data_entry = {'balance': balance, 'type': account_chart_type}

        if account_chart_type == 'Assets':
            self.financial_data['balance_sheet'][normalized_name] = account_data_entry
        elif account_chart_type == 'Liability':
            self.financial_data['balance_sheet'][normalized_name] = account_data_entry
        elif account_chart_type == 'Revenues' or account_chart_type == 'Expenses':
            self.financial_data['income_statement'][normalized_name] = account_data_entry
        elif account_chart_type == 'Equity':
            if 'retained earnings' in normalized_name:
                self.financial_data['retained_earnings']['balance_from_journal'] = account_data_entry
                logger.warning(f"Found RE balance '{balance}' from Journal for '{totals['display_name']}'.")
            self.financial_data['balance_sheet'][normalized_name] = account_data_entry
        else:
            logger.warning(f"Account '{totals['display_name']}' has unclassified chart type '{account_chart_type}'. Skipping.")
            skipped_accounts.append(totals['display_name'])

        processed_accounts_count += 1

    if skipped_accounts:
         logger.warning(f"Skipped {len(skipped_accounts)} accounts during Journal processing (not found in chart): {skipped_accounts}")
    logger.info(f"Successfully calculated balances for {processed_accounts_count} accounts from Journal sheet.")

# Tambahkan metode ke kelas AccountingChatbotGUI
AccountingChatbotGUI.load_financial_data = load_financial_data
AccountingChatbotGUI.is_trial_balance_sheet = is_trial_balance_sheet
AccountingChatbotGUI.is_journal_sheet = is_journal_sheet
AccountingChatbotGUI.load_trial_balance = load_trial_balance
AccountingChatbotGUI.load_journal_data = load_journal_data


# In[29]:


# Cell 4.14 - Generator Financial Statements

def format_currency(self, amount):
    """Format mata uang sesuai standar IFRS dengan tanda kurung untuk nilai negatif."""
    # Menggunakan atribut instance jika ada, jika tidak default ke Rupiah
    currency_choice = getattr(self, 'currency_var', tk.StringVar(value="Rupiah")).get()
    if currency_choice == "Rupiah":
        currency_symbol = "Rp"
    elif currency_choice == "US Dollar":
        currency_symbol = "US$"
    elif currency_choice == "Euro":
        currency_symbol = "€"
    else:
        currency_symbol = "" # Default jika tidak terdefinisi

    # Format angka dengan pemisah ribuan dan 2 desimal
    # Tambahkan spasi di depan nilai positif agar sejajar dengan negatif dalam kurung
    try:
        if amount < 0:
            formatted_amount = f"({currency_symbol}{abs(amount):,.2f})"
        elif amount > 0:
             formatted_amount = f" {currency_symbol}{amount:,.2f} "
        else: # amount == 0 or not a number? Handle zero case explicitly
             formatted_amount = f" {currency_symbol}0.00 "
    except TypeError:
        # Handle jika 'amount' bukan angka
        logger.warning(f"format_currency received non-numeric amount: {amount}")
        formatted_amount = " N/A " # Atau representasi error lainnya

    return formatted_amount

def generate_income_statement(self):
    """Generate Income Statement di dalam frame scrollable."""
    logger.info("Generating Income Statement...")

    # (Pemeriksaan data di awal tetap sama seperti respons sebelumnya)
    if not self.financial_data.get('income_statement'):
        messagebox.showwarning("Peringatan", "Belum ada data Laporan Laba Rugi yang dimuat/disiapkan.", parent=self.income_statement_tab)
        logger.error("Failed to generate Income Statement: financial_data['income_statement'] is empty or missing.")
        return
    if not hasattr(self.assistant, 'account_data') or not self.assistant.account_data:
        messagebox.showerror("Error", "Data Chart of Accounts tidak tersedia.", parent=self.income_statement_tab)
        logger.error("Failed to generate Income Statement: assistant.account_data is missing or empty.")
        return

    # --- PERBAIKAN: Targetkan frame scrollable ---
    # Bersihkan frame scrollable sebelumnya
    for widget in self.scrollable_income_frame.winfo_children():
        widget.destroy()
    logger.debug("Cleared previous scrollable Income Statement content.")

    # Header (ditempatkan di dalam frame scrollable)
    company_name = getattr(self, 'company_name', "PT Default Company")
    current_date = datetime.now().strftime("%d %B %Y")

    # Gunakan self.scrollable_income_frame sebagai parent
    header_frame = tk.Frame(self.scrollable_income_frame, bg='white')
    header_frame.pack(fill='x', pady=(0, 20)) # pack di frame scrollable
    tk.Label(header_frame, text=company_name, font=('Helvetica', 14, 'bold'), bg='white').pack()
    tk.Label(header_frame, text="INCOME STATEMENT", font=('Helvetica', 12, 'bold'), bg='white').pack()
    tk.Label(header_frame, text=f"For the Period Ended {current_date}", font=('Helvetica', 10), bg='white').pack()

    # Frame Tabel (juga di dalam frame scrollable)
    # Gunakan self.scrollable_income_frame sebagai parent
    table_frame = tk.Frame(self.scrollable_income_frame, bg='white')
    table_frame.pack(fill='x', expand=True, padx=20) # fill='x' agar tidak terlalu lebar
    # --- AKHIR PERBAIKAN TARGET FRAME ---

    # (Logika pengelompokan akun dan perhitungan total tetap sama)
    revenues = []
    expenses = []
    for account_name_lower, data in self.financial_data['income_statement'].items():
        balance = data.get('balance', 0)
        account_type = data.get('type', 'Unknown').strip().capitalize()
        account_info = self.assistant.account_data.get(account_name_lower, {})
        account_display_name = account_info.get('original_name', account_name_lower.title())
        if account_type == 'Revenues': revenues.append((account_display_name, balance))
        elif account_type == 'Expenses': expenses.append((account_display_name, balance))
    total_revenue = sum(amount for _, amount in revenues)
    total_expense = sum(amount for _, amount in expenses)
    net_income = total_revenue - total_expense
    if 'retained_earnings' not in self.financial_data: self.financial_data['retained_earnings'] = {}
    self.financial_data['retained_earnings']['net_income'] = {'balance': net_income, 'type': 'Equity'}


    # --- PERBAIKAN: Tempatkan widget di table_frame (yg ada di scrollable_income_frame) ---
    # Tampilkan Bagian Pendapatan
    tk.Label(table_frame, text="Revenues", font=('Helvetica', 11, 'bold'), bg='white', anchor='w').pack(fill='x', pady=(10, 5))
    if not revenues: tk.Label(table_frame, text="   (No revenue accounts found)", font=('Helvetica', 9, 'italic'), bg='white', anchor='w').pack(fill='x')
    else:
        for name, amount in sorted(revenues):
            row_frame = tk.Frame(table_frame, bg='white')
            row_frame.pack(fill='x')
            tk.Label(row_frame, text=f"   {name}", width=40, anchor='w', bg='white', font=('Helvetica', 10)).pack(side='left', padx=(10, 0))
            tk.Label(row_frame, text=self.format_currency(amount), width=20, anchor='e', bg='white', font=('Helvetica', 10)).pack(side='right')

    # Total Pendapatan
    total_rev_frame = tk.Frame(table_frame, bg='white')
    total_rev_frame.pack(fill='x', pady=(5, 15))
    tk.Frame(total_rev_frame, height=1, bg='grey').pack(fill='x', side='top', padx=(50, 0))
    tk.Label(total_rev_frame, text="Total Revenues", width=40, anchor='w', font=('Helvetica', 10, 'bold'), bg='white').pack(side='left')
    tk.Label(total_rev_frame, text=self.format_currency(total_revenue), width=20, anchor='e', font=('Helvetica', 10, 'bold'), bg='white').pack(side='right')

    # Tampilkan Bagian Beban
    tk.Label(table_frame, text="Expenses", font=('Helvetica', 11, 'bold'), bg='white', anchor='w').pack(fill='x', pady=(10, 5))
    if not expenses: tk.Label(table_frame, text="   (No expense accounts found)", font=('Helvetica', 9, 'italic'), bg='white', anchor='w').pack(fill='x')
    else:
        for name, amount in sorted(expenses):
            row_frame = tk.Frame(table_frame, bg='white')
            row_frame.pack(fill='x')
            tk.Label(row_frame, text=f"   {name}", width=40, anchor='w', bg='white', font=('Helvetica', 10)).pack(side='left', padx=(10, 0))
            tk.Label(row_frame, text=self.format_currency(amount), width=20, anchor='e', bg='white', font=('Helvetica', 10)).pack(side='right')

    # Total Beban
    total_exp_frame = tk.Frame(table_frame, bg='white')
    total_exp_frame.pack(fill='x', pady=(5, 15))
    tk.Frame(total_exp_frame, height=1, bg='grey').pack(fill='x', side='top', padx=(50, 0))
    tk.Label(total_exp_frame, text="Total Expenses", width=40, anchor='w', font=('Helvetica', 10, 'bold'), bg='white').pack(side='left')
    tk.Label(total_exp_frame, text=self.format_currency(total_expense), width=20, anchor='e', font=('Helvetica', 10, 'bold'), bg='white').pack(side='right')

    # Laba Bersih
    separator = tk.Frame(table_frame, height=2, bg='black')
    separator.pack(fill='x', pady=5)
    net_frame = tk.Frame(table_frame, bg='white')
    net_frame.pack(fill='x', pady=5)
    tk.Label(net_frame, text="Net Income" if net_income >= 0 else "Net Loss", width=40, anchor='w', font=('Helvetica', 11, 'bold'), bg='white').pack(side='left')
    tk.Label(net_frame, text=self.format_currency(net_income), width=20, anchor='e', font=('Helvetica', 11, 'bold'), bg='white').pack(side='right')
    # --- AKHIR PERBAIKAN TARGET FRAME ---

    self.financial_status_var.set(f"Status: Laporan Laba Rugi berhasil digenerate")
    logger.info("Income Statement generation complete.")

def generate_retained_earnings(self):
    """Generate Retained Earnings Statement dengan penanganan dividen yang benar."""
    logger.info("Generating Retained Earnings Statement...")

    re_data = self.financial_data.get('retained_earnings', {})

    # Cek Net Income
    if 'net_income' not in re_data:
        logger.warning("Net Income missing. Generating Income Statement first...")
        self.generate_income_statement()
        re_data = self.financial_data.get('retained_earnings', {})
        if 'net_income' not in re_data:
             messagebox.showwarning("Peringatan", "Data Laba Bersih tidak ditemukan.", parent=self.retained_earnings_tab)
             logger.error("Failed to generate RE Statement: Net Income missing.")
             return
    net_income = re_data.get('net_income', {}).get('balance', 0)

    # Cek Beginning Balance 
    beginning_balance_data = re_data.get('beginning_balance', {'balance': 0})
    beginning_balance = beginning_balance_data.get('balance', 0)
    logger.debug(f"Using Beginning Retained Earnings: {beginning_balance}")

    # --- PERBAIKAN: Penanganan Saldo Dividen ---
    dividends_balance_raw = 0 # Saldo asli dari perhitungan (bisa negatif)
    dividends_account_found = False
    # Cari akun yang mengandung 'dividends' (lebih fleksibel)
    for acc_lower, data in self.financial_data.get('balance_sheet', {}).items():
        if 'dividends' in acc_lower: # Cari kata 'dividends'
            account_info = self.assistant.account_data.get(acc_lower, {})
            # Pastikan itu akun Equity dengan normal Debit (kontra ekuitas)
            if account_info.get('account_chart', '').strip().capitalize() == 'Equity' and \
               account_info.get('normal_account_position', '').strip().capitalize() == 'Debit':
                dividends_balance_raw = data.get('balance', 0)
                dividends_account_found = True
                logger.debug(f"Found Dividends account '{acc_lower}' with raw balance: {dividends_balance_raw}")
                break # Asumsi hanya ada satu akun dividen utama
    # Dividen mengurangi RE, jadi kita butuh nilai absolutnya jika saldo dihitung negatif,
    # atau nilai positifnya jika saldo dihitung positif karena cara perhitungan saldo.
    # Karena dividen normalnya Debit, perhitungan (Kredit - Debit) akan menghasilkan negatif.
    # Jadi, kita gunakan nilai absolutnya sebagai pengurang.
    dividends_subtraction_amount = abs(dividends_balance_raw)
    logger.debug(f"Amount to subtract for dividends: {dividends_subtraction_amount}")
    # Simpan jumlah pengurang untuk konsistensi jika diperlukan di tempat lain
    self.financial_data['retained_earnings']['dividends'] = {
        'balance': dividends_subtraction_amount, 'type': 'Equity' # Tetap tipe Equity
    }
    # --- AKHIR PERBAIKAN Dividen ---

    # Bersihkan frame scrollable
    for widget in self.scrollable_retained_frame.winfo_children():
        widget.destroy()

    # Header (di dalam frame scrollable)
    company_name = getattr(self, 'company_name', "PT Default Company")
    current_date = datetime.now().strftime("%d %B %Y")
    header_frame = tk.Frame(self.scrollable_retained_frame, bg='white')
    header_frame.pack(fill='x', pady=(0, 20))
    tk.Label(header_frame, text=company_name, font=('Helvetica', 14, 'bold'), bg='white').pack()
    tk.Label(header_frame, text="STATEMENT OF RETAINED EARNINGS", font=('Helvetica', 12, 'bold'), bg='white').pack()
    tk.Label(header_frame, text=f"For the Period Ended {current_date}", font=('Helvetica', 10), bg='white').pack()

    # Frame Tabel (di dalam frame scrollable)
    table_frame = tk.Frame(self.scrollable_retained_frame, bg='white')
    table_frame.pack(fill='x', expand=True, padx=20)

    # Hitung Saldo Akhir (menggunakan nilai absolut dividen sebagai pengurang)
    ending_balance = beginning_balance + net_income - dividends_subtraction_amount
    logger.info(f"Calculated Ending RE: {ending_balance} (Begin: {beginning_balance}, Net Income: {net_income}, Dividends Subtracted: {dividends_subtraction_amount})")
    self.financial_data['retained_earnings']['ending_balance'] = {'balance': ending_balance, 'type': 'Equity'}

    # Tampilkan Baris-baris (menggunakan nilai absolut dividen sebagai pengurang)
    # Saldo Awal
    row_frame = tk.Frame(table_frame, bg='white'); row_frame.pack(fill='x', pady=5)
    tk.Label(row_frame, text="Retained Earnings, Beginning", width=40, anchor='w', font=('Helvetica', 10), bg='white').pack(side='left')
    tk.Label(row_frame, text=self.format_currency(beginning_balance), width=20, anchor='e', font=('Helvetica', 10), bg='white').pack(side='right')
    # Tambah: Laba Bersih
    row_frame = tk.Frame(table_frame, bg='white'); row_frame.pack(fill='x', pady=5)
    tk.Label(row_frame, text="Add: Net Income" if net_income >= 0 else "Add: Net Loss", width=40, anchor='w', font=('Helvetica', 10), bg='white').pack(side='left')
    tk.Label(row_frame, text=self.format_currency(net_income), width=20, anchor='e', font=('Helvetica', 10), bg='white').pack(side='right')
    # Subtotal
    subtotal = beginning_balance + net_income
    row_frame = tk.Frame(table_frame, bg='white'); row_frame.pack(fill='x', pady=5)
    tk.Frame(row_frame, height=1, bg='grey').pack(fill='x', side='top', padx=(200, 0))
    tk.Label(row_frame, text="Subtotal", width=40, anchor='w', font=('Helvetica', 10, 'bold'), bg='white').pack(side='left')
    tk.Label(row_frame, text=self.format_currency(subtotal), width=20, anchor='e', font=('Helvetica', 10, 'bold'), bg='white').pack(side='right')
    # Kurang: Dividen
    row_frame = tk.Frame(table_frame, bg='white'); row_frame.pack(fill='x', pady=5)
    tk.Label(row_frame, text="Less: Dividends", width=40, anchor='w', font=('Helvetica', 10), bg='white').pack(side='left')
    tk.Label(row_frame, text=self.format_currency(dividends_subtraction_amount), width=20, anchor='e', font=('Helvetica', 10), bg='white').pack(side='right') # Tampilkan nilai positif pengurang
    # Separator
    separator = tk.Frame(table_frame, height=2, bg='black'); separator.pack(fill='x', pady=5)
    # Saldo Akhir
    row_frame = tk.Frame(table_frame, bg='white'); row_frame.pack(fill='x', pady=5)
    tk.Label(row_frame, text="Retained Earnings, Ending", width=40, anchor='w', font=('Helvetica', 11, 'bold'), bg='white').pack(side='left')
    tk.Label(row_frame, text=self.format_currency(ending_balance), width=20, anchor='e', font=('Helvetica', 11, 'bold'), bg='white').pack(side='right')

    self.financial_status_var.set(f"Status: Laporan Laba Ditahan berhasil digenerate")
    logger.info("Retained Earnings Statement generation complete.")

def generate_balance_sheet(self):
    """Generate Balance Sheet di dalam frame scrollable."""
    logger.info("Generating Balance Sheet...")

    # (Pemeriksaan data di awal tetap sama seperti respons sebelumnya) 
    re_data = self.financial_data.get('retained_earnings', {})
    if 'ending_balance' not in re_data:
        logger.warning("Ending RE missing. Attempting RE generation...")
        self.generate_retained_earnings()
        re_data = self.financial_data.get('retained_earnings', {})
        if 'ending_balance' not in re_data:
            messagebox.showwarning("Peringatan", "Data Laba Ditahan Akhir tidak ditemukan.", parent=self.balance_sheet_tab)
            logger.error("Failed to generate BS: Ending RE data missing.")
            return
    if not self.financial_data.get('balance_sheet'):
         messagebox.showwarning("Peringatan", "Data Akun Neraca tidak ditemukan.", parent=self.balance_sheet_tab)
         logger.error("Failed to generate BS: financial_data['balance_sheet'] missing.")
         return
    if not hasattr(self.assistant, 'account_data') or not self.assistant.account_data:
        messagebox.showerror("Error", "Data Chart of Accounts tidak tersedia.", parent=self.balance_sheet_tab)
        logger.error("Failed to generate BS: assistant.account_data missing.")
        return

    # --- PERBAIKAN: Targetkan frame scrollable ---
    # Bersihkan frame scrollable sebelumnya
    for widget in self.scrollable_balance_frame.winfo_children():
        widget.destroy() 
    logger.debug("Cleared previous scrollable Balance Sheet content.")

    # Header (di dalam frame scrollable)
    company_name = getattr(self, 'company_name', "PT Default Company")
    current_date = datetime.now().strftime("%d %B %Y")
    # Gunakan self.scrollable_balance_frame sebagai parent
    header_frame = tk.Frame(self.scrollable_balance_frame, bg='white')
    header_frame.pack(fill='x', pady=(0, 20)) # pack di frame scrollable
    tk.Label(header_frame, text=company_name, font=('Helvetica', 14, 'bold'), bg='white').pack()
    tk.Label(header_frame, text="STATEMENT OF FINANCIAL POSITION", font=('Helvetica', 12, 'bold'), bg='white').pack()
    tk.Label(header_frame, text=f"As of {current_date}", font=('Helvetica', 10), bg='white').pack()

    # Frame Tabel (di dalam frame scrollable)
    # Gunakan self.scrollable_balance_frame sebagai parent
    table_frame = tk.Frame(self.scrollable_balance_frame, bg='white')
    table_frame.pack(fill='x', expand=True, padx=20) # fill='x'
    # --- AKHIR PERBAIKAN TARGET FRAME ---


    # (Logika pengelompokan akun dan perhitungan total tetap sama)
    assets, liabilities, equities, unknown_accounts = [], [], [], []
    for account_name_lower, data in self.financial_data['balance_sheet'].items():
        balance = data.get('balance', 0)
        account_info = self.assistant.account_data.get(account_name_lower)  
        if account_info:
            account_chart_type = account_info.get('account_chart', 'Unknown').strip().capitalize()
            account_display_name = account_info.get('original_name', account_name_lower.title())
            if 'retained earnings' in account_name_lower: continue # Skip RE here
            if account_chart_type == 'Assets': assets.append((account_display_name, balance))
            elif account_chart_type == 'Liability': liabilities.append((account_display_name, balance))
            elif account_chart_type == 'Equity': equities.append((account_display_name, balance))
            else: unknown_accounts.append((account_display_name, balance))
        else: unknown_accounts.append((account_name_lower.title(), balance))
    retained_earnings_ending_balance = re_data.get('ending_balance', {}).get('balance', 0)
    equities.append(("Retained Earnings", retained_earnings_ending_balance))
    total_assets = sum(amount for _, amount in assets)
    total_liabilities = sum(amount for _, amount in liabilities)
    total_equities = sum(amount for _, amount in equities)
    total_liabilities_equity = total_liabilities + total_equities


    # --- PERBAIKAN: Modifikasi add_section untuk menerima parent frame ---
    # Definisikan helper function di dalam generate_balance_sheet
    def add_section(parent_frame, title, items):
        # Gunakan parent_frame yang diberikan, bukan table_frame global
        tk.Label(parent_frame, text=title, font=('Helvetica', 11, 'bold'),
                 bg='white', anchor='w').pack(fill='x', pady=(10, 5))
        if not items:
             tk.Label(parent_frame, text="   (No accounts in this category)", font=('Helvetica', 9, 'italic'),
                      bg='white', anchor='w').pack(fill='x')
             return 0

        section_total = 0
        for name, amount in sorted(items):
            row_frame = tk.Frame(parent_frame, bg='white') # Parent adalah parent_frame
            row_frame.pack(fill='x')
            tk.Label(row_frame, text=f"   {name}", width=40, anchor='w', bg='white', font=('Helvetica', 10)).pack(side='left', padx=(10,0))
            # Panggil self.format_currency
            tk.Label(row_frame, text=self.format_currency(amount), width=20, anchor='e', bg='white', font=('Helvetica', 10)).pack(side='right')
            section_total += amount

        total_frame = tk.Frame(parent_frame, bg='white') # Parent adalah parent_frame
        total_frame.pack(fill='x', pady=(5, 15))
        tk.Frame(total_frame, height=1, bg='grey').pack(fill='x', side='top', padx=(50, 0))
        tk.Label(total_frame, text=f"Total {title}", width=40, anchor='w',
                 font=('Helvetica', 10, 'bold'), bg='white').pack(side='left')
        tk.Label(total_frame, text=self.format_currency(section_total), width=20, anchor='e',
                 font=('Helvetica', 10, 'bold'), bg='white').pack(side='right')
        return section_total
    # --- AKHIR PERBAIKAN add_section ---


    # --- PERBAIKAN: Panggil add_section dengan parent frame yang benar ---
    # Panggil add_section dengan table_frame (yang ada di scrollable_balance_frame)
    actual_total_assets = add_section(table_frame, "Assets", assets)
    actual_total_liabilities = add_section(table_frame, "Liabilities", liabilities)  
    actual_total_equity = add_section(table_frame, "Equity", equities)
    if unknown_accounts:
         add_section(table_frame, "Unknown/Unclassified Accounts", unknown_accounts)

    # Total Liabilities and Equity (parent = table_frame)
    separator = tk.Frame(table_frame, height=2, bg='black')
    separator.pack(fill='x', pady=5)
    total_le_frame = tk.Frame(table_frame, bg='white')
    total_le_frame.pack(fill='x', pady=5)
    tk.Label(total_le_frame, text="Total Liabilities and Equity", width=40, anchor='w',
             font=('Helvetica', 11, 'bold'), bg='white').pack(side='left')
    tk.Label(total_le_frame, text=self.format_currency(total_liabilities_equity), width=20, anchor='e',
             font=('Helvetica', 11, 'bold'), bg='white').pack(side='right')

    # Pemeriksaan Keseimbangan (parent = table_frame)
    balance_check_frame = tk.Frame(table_frame, bg='white')
    balance_check_frame.pack(fill='x', pady=(10, 5))
    if abs(total_assets - total_liabilities_equity) < 0.01:
        check_text = "Assets = Liabilities + Equity"  
        check_color = 'green'
    else:
        diff = total_assets - total_liabilities_equity
        check_text = f"UNBALANCED! Difference: {self.format_currency(diff)}"
        check_color = 'red'
    tk.Label(balance_check_frame, text=check_text, font=('Helvetica', 10, 'bold'), fg=check_color, bg='white').pack()
    # --- AKHIR PERBAIKAN PEMANGGILAN add_section ---


    self.financial_status_var.set(f"Status: Neraca berhasil digenerate")
    logger.info("Balance Sheet generation complete.")

# Add the format_currency method to the class if it's not already there
if not hasattr(AccountingChatbotGUI, 'format_currency'):
    AccountingChatbotGUI.format_currency = format_currency
    
# Tambahkan metode ke kelas AccountingChatbotGUI
AccountingChatbotGUI.generate_income_statement = generate_income_statement
AccountingChatbotGUI.generate_retained_earnings = generate_retained_earnings
AccountingChatbotGUI.generate_balance_sheet = generate_balance_sheet


# In[30]:


# Cell 4.15 - Export Financial Statements to Excel

def export_statement_to_excel(self, statement_type):
    """Export financial statement ke Excel"""
    if statement_type == 'income' and not self.financial_data['income_statement']:
        messagebox.showwarning("Peringatan", "Belum ada data Income Statement. Silakan generate terlebih dahulu.")
        return
    elif statement_type == 'retained' and 'ending_balance' not in self.financial_data['retained_earnings']:
        messagebox.showwarning("Peringatan", "Belum ada data Retained Earnings. Silakan generate terlebih dahulu.")
        return
    elif statement_type == 'balance' and not self.financial_data['balance_sheet']:
        messagebox.showwarning("Peringatan", "Belum ada data Balance Sheet. Silakan generate terlebih dahulu.")
        return
    
    # Get current date for filename
    current_date = datetime.now().strftime("%Y-%m-%d")
    
    # Set default filename based on statement type
    if statement_type == 'income':
        default_filename = f"Income_Statement_{current_date}.xlsx"
        title = "Income Statement"
    elif statement_type == 'retained':
        default_filename = f"Retained_Earnings_{current_date}.xlsx"
        title = "Statement of Retained Earnings"
    else:  # balance
        default_filename = f"Balance_Sheet_{current_date}.xlsx"
        title = "Statement of Financial Position"
    
    # Ask for save location
    filename = filedialog.asksaveasfilename(
        initialfile=default_filename,
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
    )
    
    if not filename:
        return  # User cancelled
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = title.replace(" ", "_")
    
    # Set company name and date
    company_name = "AI Accounting Assistant Company"
    date_str = datetime.now().strftime("%d %B %Y")
    
    # Add header
    ws.merge_cells('A1:C1')
    ws['A1'] = company_name
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:C2')
    ws['A2'] = title
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:C3')
    if statement_type in ['income', 'retained']:
        ws['A3'] = f"For the Period Ended {date_str}"
    else:
        ws['A3'] = f"As of {date_str}"
    ws['A3'].font = Font(size=10)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    # Add content based on statement type
    if statement_type == 'income':
        self.export_income_statement_content(ws)
    elif statement_type == 'retained':
        self.export_retained_earnings_content(ws)
    else:  # balance
        self.export_balance_sheet_content(ws)
    
    # Save workbook
    try:
        wb.save(filename)
        messagebox.showinfo("Sukses", f"{title} berhasil diekspor ke {filename}")
    except Exception as e:
        messagebox.showerror("Error", f"Gagal menyimpan file: {str(e)}")

def export_income_statement_content(self, ws):
    """Ekspor konten Income Statement ke worksheet"""
    row = 5  # Start after header
    
    # Calculate totals
    revenues = [(name, data['balance']) for name, data in self.financial_data['income_statement'].items() 
               if data['type'] == 'revenue']
    expenses = [(name, data['balance']) for name, data in self.financial_data['income_statement'].items() 
               if data['type'] == 'expense']
    
    total_revenue = sum(amount for _, amount in revenues)
    total_expense = sum(amount for _, amount in expenses)
    net_income = total_revenue - total_expense
    
    # Revenue section
    ws.cell(row=row, column=1, value="Revenues")
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    for name, amount in sorted(revenues):
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=3, value=amount)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        row += 1
    
    # Total revenue
    ws.cell(row=row, column=1, value="Total Revenues")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_revenue)
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=3).font = Font(bold=True)
    row += 2
    
    # Expenses section
    ws.cell(row=row, column=1, value="Expenses")
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    for name, amount in sorted(expenses):
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=3, value=amount)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        row += 1
    
    # Total expenses
    ws.cell(row=row, column=1, value="Total Expenses")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_expense)
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=3).font = Font(bold=True)
    row += 2
    
    # Net Income
    for col in range(1, 4):
        cell = ws.cell(row=row, column=col)
        cell.border = Border(top=Side(style='thin'))
    row += 1
    
    ws.cell(row=row, column=1, value="Net Income")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=net_income)
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=3).font = Font(bold=True)

def export_retained_earnings_content(self, ws):
    """Ekspor konten Retained Earnings Statement ke worksheet"""
    row = 5  # Start after header
    
    # Get values
    beginning_balance = self.financial_data['retained_earnings']['beginning_balance']['balance']
    net_income = self.financial_data['retained_earnings']['net_income']['balance']
    dividends = self.financial_data['retained_earnings']['dividends']['balance']
    ending_balance = self.financial_data['retained_earnings']['ending_balance']['balance']
    
    # Beginning Balance
    ws.cell(row=row, column=1, value="Retained Earnings, Beginning")
    ws.cell(row=row, column=3, value=beginning_balance)
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    row += 1
    
    # Add Net Income
    ws.cell(row=row, column=1, value="Add: Net Income")
    ws.cell(row=row, column=3, value=net_income)
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    row += 1
    
    # Subtotal
    subtotal = beginning_balance + net_income
    ws.cell(row=row, column=1, value="Subtotal")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=subtotal)
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=3).font = Font(bold=True)
    row += 1
    
    # Less: Dividends
    ws.cell(row=row, column=1, value="Less: Dividends")
    ws.cell(row=row, column=3, value=dividends)
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    row += 1
    
    # Separator
    for col in range(1, 4):
        cell = ws.cell(row=row, column=col)
        cell.border = Border(top=Side(style='thin'))
    row += 1
    
    # Ending Balance
    ws.cell(row=row, column=1, value="Retained Earnings, Ending")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=ending_balance)
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=3).font = Font(bold=True)

def export_balance_sheet_content(self, ws):
    """Ekspor konten Balance Sheet ke worksheet"""
    row = 5  # Start after header
    
    # Group accounts by type
    assets = [(name, data['balance']) for name, data in self.financial_data['balance_sheet'].items() 
             if data['type'] == 'asset']
    liabilities = [(name, data['balance']) for name, data in self.financial_data['balance_sheet'].items() 
                 if data['type'] == 'liability']
    equities = [(name, data['balance']) for name, data in self.financial_data['balance_sheet'].items() 
              if data['type'] == 'equity' and 'retained earnings' not in name.lower()]
    
    # Add retained earnings from retained earnings statement
    retained_earnings = self.financial_data['retained_earnings']['ending_balance']['balance']
    equities.append(("Retained Earnings", retained_earnings))
    
    # Calculate totals
    total_assets = sum(amount for _, amount in assets)
    total_liabilities = sum(amount for _, amount in liabilities)
    total_equities = sum(amount for _, amount in equities)
    total_liabilities_equity = total_liabilities + total_equities
    
    # Assets section
    ws.cell(row=row, column=1, value="Assets")
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    for name, amount in sorted(assets):
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=3, value=amount)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        row += 1
    
    # Total assets
    ws.cell(row=row, column=1, value="Total Assets")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_assets)
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=3).font = Font(bold=True)
    row += 2
    
    # Liabilities section
    ws.cell(row=row, column=1, value="Liabilities")
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    for name, amount in sorted(liabilities):
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=3, value=amount)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        row += 1
    
    # Total liabilities
    ws.cell(row=row, column=1, value="Total Liabilities")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_liabilities)
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=3).font = Font(bold=True)
    row += 2
    
    # Equity section
    ws.cell(row=row, column=1, value="Equity")
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    for name, amount in sorted(equities):
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=3, value=amount)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        row += 1
    
    # Total equity
    ws.cell(row=row, column=1, value="Total Equity")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_equities)
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=3).font = Font(bold=True)
    row += 2
    
    # Total liabilities and equity
    for col in range(1, 4):
        cell = ws.cell(row=row, column=col)
        cell.border = Border(top=Side(style='thin'))
    row += 1
    
    ws.cell(row=row, column=1, value="Total Liabilities and Equity")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_liabilities_equity)
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=3).font = Font(bold=True)

def reset_journal_tab(self):
    """Mereset state tab Prediksi Jurnal."""
    logger.info("Mereset Tab Prediksi Jurnal...")
    # Kosongkan input fields
    self.input_var.set("")
    self.nominal_var.set("")
    self.currency_var.set("Rupiah") # Kembali ke default

    # Kosongkan history area dan tampilkan pesan awal
    self.history_area.configure(state='normal')
    self.history_area.delete("1.0", tk.END)
    for sender, message in self.initial_journal_messages:
        self.history_area.insert(tk.END, f"{sender}: {message}\n\n")
    self.history_area.configure(state='disabled')
    self.history_area.see(tk.END)

    # Kosongkan antrian transaksi yang diproses
    self.processed_journal_entries = []
    logger.info("Antrian transaksi untuk export buku besar telah dikosongkan.")

    messagebox.showinfo("Reset", "Tab Prediksi Jurnal telah direset.", parent=self.journal_tab)

def reset_ledger_tab(self):
    """Mereset state tab Buku Besar (termasuk data finansial yang bergantung padanya)."""
    logger.info("Mereset Tab Buku Besar...")
    # Kosongkan path file
    self.file_path.set("")

    # Hapus data ledger dan data mentah
    self.ledger_data = {}
    self.raw_journal_data = []

    # Hapus jurnal penyesuaian
    self.adjusting_entries = []
    self.update_adjustment_status() # Update label status AJE

    # Bersihkan tombol akun
    for widget in self.account_buttons_frame.winfo_children():
        widget.destroy()

    # Bersihkan tampilan buku besar dan tampilkan pesan awal
    for widget in self.ledger_content_frame.winfo_children():
        widget.destroy()
    self.guide_label = tk.Label(self.ledger_content_frame, text=self.initial_ledger_message,
                              font=self.style['title_font'], bg='white')
    self.guide_label.pack(pady=50)
    self.ledger_canvas.configure(scrollregion=self.ledger_canvas.bbox("all")) # Update scroll region

    # --- SINKRONISASI: Reset juga tab Financial Statement ---
    logger.info("Mereset Tab Financial Statement karena Buku Besar direset.")
    self.reset_financial_tab(show_message=False) # show_message=False agar tidak muncul 2 messagebox
    # --- End Sinkronisasi ---

    messagebox.showinfo("Reset", "Tab Buku Besar (dan Financial Statement) telah direset.", parent=self.ledger_tab)

def reset_financial_tab(self, show_message=True):
    """Mereset state tab Financial Statement."""
    logger.info("Mereset Tab Financial Statement...")
    # Kosongkan path file
    self.financial_file_path.set("")

    # Reset data finansial
    self.financial_data = {
        'income_statement': {},
        'retained_earnings': {},
        'balance_sheet': {}
    }

    # Reset label status
    self.financial_status_var.set("Status: Belum ada data yang dimuat")

    # Bersihkan konten sub-tab dan tampilkan pesan default
    for frame, label_attr, message in [
        (self.income_content_frame, 'income_default_label', "Belum ada data Income Statement. Silakan generate."),
        (self.retained_content_frame, 'retained_default_label', "Belum ada data Retained Earnings. Silakan generate."),
        (self.balance_content_frame, 'balance_default_label', "Belum ada data Balance Sheet. Silakan generate.")
    ]:
        for widget in frame.winfo_children():
            widget.destroy()
        default_label = tk.Label(frame, text=message, font=self.style['font'], bg='white')
        setattr(self, label_attr, default_label) # Simpan referensi jika perlu
        default_label.pack(pady=50)

    if show_message:
        messagebox.showinfo("Reset", "Tab Financial Statement telah direset.", parent=self.financial_tab)

# Tambahkan metode ke kelas AccountingChatbotGUI
AccountingChatbotGUI.export_statement_to_excel = export_statement_to_excel
AccountingChatbotGUI.export_income_statement_content = export_income_statement_content
AccountingChatbotGUI.export_retained_earnings_content = export_retained_earnings_content
AccountingChatbotGUI.export_balance_sheet_content = export_balance_sheet_content
AccountingChatbotGUI.reset_journal_tab = reset_journal_tab
AccountingChatbotGUI.reset_ledger_tab = reset_ledger_tab
AccountingChatbotGUI.reset_financial_tab = reset_financial_tab


# In[31]:


# Cell 5 - Main Execution
if __name__ == "__main__":
    login = LoginWindow()
    success, company_name = login.run()  # Dapatkan status login dan nama perusahaan

    if success:
        logger.info(f"Login berhasil untuk perusahaan: {company_name}")
        try:
            # Buat instance Assistant SETELAH login berhasil
            assistant = AccountingAssistant()

            # Buat instance GUI SETELAH assistant dibuat, teruskan assistant & company_name
            logger.info("Memulai GUI Accounting Chatbot...")
            app = AccountingChatbotGUI(assistant, company_name)  # Teruskan assistant & company_name
            logger.info("Aplikasi GUI ditutup.")

        except Exception as e:
            logger.error(f"Gagal memulai aplikasi utama: {e}", exc_info=True)
            messagebox.showerror("Error Kritis", f"Gagal memulai aplikasi: {e}")
    else:
        logger.info("Login dibatalkan atau gagal.")
